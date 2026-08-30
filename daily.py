from flask import Flask, render_template, request, redirect, url_for, flash, session, g
import mysql.connector
import os
import re
import secrets
import urllib.parse
from functools import wraps
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from flask import send_file
from openpyxl.cell.cell import MergedCell

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

# Upload folder absolute path
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# ====================
# USER AGENT PARSER (Tanpa library tambahan)
# ====================
def parse_user_agent(ua_string):
    """Parse User-Agent string untuk extract device, browser, OS."""
    if not ua_string:
        return {'device_type': 'Unknown', 'browser': 'Unknown', 'os': 'Unknown'}

    ua = ua_string.lower()

    # ── Device Type ──
    if any(k in ua for k in ['mobile', 'android', 'iphone', 'ipod', 'windows phone', 'blackberry']):
        device_type = 'Mobile'
    elif any(k in ua for k in ['tablet', 'ipad', 'kindle', 'silk']):
        device_type = 'Tablet'
    else:
        device_type = 'Desktop'

    # ── Operating System ──
    os_name = 'Unknown'
    if 'windows nt 10' in ua or 'windows 10' in ua:
        os_name = 'Windows 10/11'
    elif 'windows nt 6.3' in ua:
        os_name = 'Windows 8.1'
    elif 'windows nt 6.2' in ua:
        os_name = 'Windows 8'
    elif 'windows nt 6.1' in ua:
        os_name = 'Windows 7'
    elif 'windows' in ua:
        os_name = 'Windows'
    elif 'iphone' in ua:
        os_name = 'iOS (iPhone)'
    elif 'ipad' in ua:
        os_name = 'iOS (iPad)'
    elif 'mac os x' in ua or 'macintosh' in ua:
        ver = re.search(r'mac os x[\s_](\d+[._]\d+)', ua)
        os_name = f'macOS {ver.group(1).replace("_", ".")}' if ver else 'macOS'
    elif 'android' in ua:
        ver = re.search(r'android[\s_](\d+[.\d]*)', ua)
        os_name = f'Android {ver.group(1)}' if ver else 'Android'
    elif 'linux' in ua:
        os_name = 'Linux'
    elif 'cros' in ua:
        os_name = 'Chrome OS'

    # ── Browser ──
    browser = 'Unknown'
    if 'edg/' in ua or 'edge/' in ua:
        ver = re.search(r'edg[e/][\s/](\d+[.\d]*)', ua)
        browser = f'Microsoft Edge {ver.group(1)}' if ver else 'Microsoft Edge'
    elif 'opr/' in ua or 'opera' in ua:
        ver = re.search(r'opr/[\s/](\d+[.\d]*)', ua)
        browser = f'Opera {ver.group(1)}' if ver else 'Opera'
    elif 'vivaldi' in ua:
        ver = re.search(r'vivaldi/[\s/](\d+[.\d]*)', ua)
        browser = f'Vivaldi {ver.group(1)}' if ver else 'Vivaldi'
    elif 'brave' in ua:
        browser = 'Brave'
    elif 'chrome' in ua and 'safari' in ua:
        ver = re.search(r'chrome/[\s/](\d+[.\d]*)', ua)
        browser = f'Google Chrome {ver.group(1)}' if ver else 'Google Chrome'
    elif 'firefox' in ua:
        ver = re.search(r'firefox/[\s/](\d+[.\d]*)', ua)
        browser = f'Mozilla Firefox {ver.group(1)}' if ver else 'Mozilla Firefox'
    elif 'safari' in ua and 'chrome' not in ua:
        ver = re.search(r'version/[\s/](\d+[.\d]*)', ua)
        browser = f'Safari {ver.group(1)}' if ver else 'Safari'

    return {'device_type': device_type, 'browser': browser, 'os': os_name}


# Upload tanpa pembatasan - semua format dan ukuran diterima

def allowed_file(filename, allowed=None):
    # Selalu return True - tidak ada pembatasan format
    return '.' in filename

def validate_upload(file, allowed=None, label='File'):
    # Selalu valid - tidak ada pembatasan ukuran atau format
    if not file or file.filename == '':
        return True, ''
    return True, ''

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            flash('Silakan login terlebih dahulu!', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def generate_csrf_token():
    if '_csrf_token' not in session:
        session['_csrf_token'] = secrets.token_hex(32)
    return session['_csrf_token']

app.jinja_env.globals['csrf_token'] = generate_csrf_token

def validate_csrf():
    token = session.pop('_csrf_token', None)
    if not token or token != request.form.get('_csrf_token'):
        flash('Invalid CSRF token. Silakan coba lagi.', 'danger')
        return False
    return True

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not session.get('logged_in'):
                flash('Silakan login terlebih dahulu!', 'warning')
                return redirect(url_for('login'))
            if session.get('role', 'user') not in roles:
                flash(f'Akses ditolak. Hanya role {"/".join(roles)} yang diizinkan.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def get_db_connection():
    """Koneksi database - support MySQL lokal maupun cloud (Railway, TiDB, dll)."""
    import urllib.parse
    
    # Cek semua kemungkinan env var untuk database URL
    database_url = (
        os.environ.get('DATABASE_URL', '') or 
        os.environ.get('MYSQL_URL', '') or 
        os.environ.get('MYSQLURL', '')
    )
    
    # Default values
    db_host = os.environ.get('MYSQLHOST', '') or os.environ.get('DB_HOST', '127.0.0.1')
    db_port = int(os.environ.get('MYSQLPORT', '') or os.environ.get('DB_PORT', '3306') or '3306')
    db_user = os.environ.get('MYSQLUSER', '') or os.environ.get('DB_USER', 'root')
    db_password = os.environ.get('MYSQLPASSWORD', '') or os.environ.get('DB_PASSWORD', '')
    db_name = os.environ.get('MYSQLDATABASE', '') or os.environ.get('DB_NAME', 'db_proyek')
    
    if database_url:
        try:
            # Handle postgres:// format
            if database_url.startswith('postgres://'):
                database_url = database_url.replace('postgres://', 'mysql://', 1)
            
            # Parse URL menggunakan urllib
            if database_url.startswith('mysql://'):
                parsed = urllib.parse.urlparse(database_url)
                db_host = parsed.hostname or '127.0.0.1'
                db_port = parsed.port or 3306
                db_user = parsed.username or 'root'
                db_password = parsed.password or ''
                db_name = (parsed.path or '/railway').lstrip('/') or 'railway'
                
                # Handle URL-encoded password
                if db_password:
                    db_password = urllib.parse.unquote(db_password)
                    
                print(f'[DB] Parsed URL: host={db_host}, port={db_port}, user={db_user}, db={db_name}')
        except Exception as e:
            print(f'[DB] Error parsing URL: {e}')
    else:
        print(f'[DB] Using env vars: host={db_host}, port={db_port}, user={db_user}, db={db_name}')
    
    # Debug: print password length (bukan password-nya untuk keamanan)
    print(f'[DB] Connecting to: {db_host}:{db_port}/{db_name} as {db_user} (pwd_len={len(db_password)})')
    
    # SSL config untuk TiDB Cloud
    ssl_config = {
        'ssl_disabled': False,
        'ssl_ca': None,
        'ssl_verify_cert': False,
        'ssl_verify_identity': False
    }
    
    return mysql.connector.connect(
        host=db_host,
        port=db_port,
        user=db_user,
        password=db_password,
        database=db_name,
        **ssl_config
    )

# ====================
# HAK AKSES (MENU PERMISSIONS)
# ====================
def load_permissions(role):
    """Load menu permissions dari DB ke dict. Admin selalu punya akses penuh."""
    if role == 'admin':
        return {k: True for k in ['dashboard','pengaturan','users','active_sessions','daily_report','weekly_report','monthly_report','gantt','wbs','budget','sync']}
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT menu_key, can_access FROM menu_permissions WHERE role = %s", (role,))
        rows = cursor.fetchall()
        conn.close()
        return {r['menu_key']: bool(r['can_access']) for r in rows}
    except Exception:
        # Fallback jika tabel belum ada
        defaults = {
            'manager': {'dashboard':1,'pengaturan':1,'users':0,'active_sessions':0,'daily_report':1,'weekly_report':1,'monthly_report':1,'gantt':1,'wbs':1,'budget':1},
            'user':    {'dashboard':1,'pengaturan':0,'users':0,'active_sessions':0,'daily_report':1,'weekly_report':1,'monthly_report':1,'gantt':1,'wbs':0,'budget':0}
        }
        return {k: bool(v) for k, v in defaults.get(role, defaults['user']).items()}

def has_menu_access(menu_key):
    """Cek apakah user saat ini boleh mengakses menu tertentu."""
    perms = session.get('menu_permissions', {})
    if not perms:
        # Lazy-load jika session kosong
        role = session.get('role', 'user')
        perms = load_permissions(role)
        session['menu_permissions'] = perms
    return perms.get(menu_key, False)

app.jinja_env.globals['has_menu_access'] = has_menu_access

# ====================
# AUTO-MIGRATION: Buat tabel jika belum ada
# ====================
def auto_migrate():
    """Buat SEMUA tabel dari schema.sql + tabel baru + seed default jika belum ada."""
    SQL_SCHEMA = """
CREATE TABLE IF NOT EXISTS `users` (
    `id`           INT AUTO_INCREMENT PRIMARY KEY,
    `nama_lengkap` VARCHAR(150) NOT NULL,
    `username`     VARCHAR(80)  NOT NULL UNIQUE,
    `password_hash` VARCHAR(255) NOT NULL,
    `role`         ENUM('admin','manager','user') NOT NULL DEFAULT 'user',
    `status`       ENUM('pending','approved','rejected') NOT NULL DEFAULT 'approved',
    `created_at`   DATETIME DEFAULT CURRENT_TIMESTAMP
) ENGINE=InnoDB;

CREATE TABLE IF NOT EXISTS `master_proyek` (
    `id`                  INT AUTO_INCREMENT PRIMARY KEY,
    `nama_kegiatan`       VARCHAR(255) NOT NULL,
    `nama_proyek`         VARCHAR(255) NOT NULL,
    `penyedia_jasa`       VARCHAR(255) DEFAULT NULL,
    `konsultan`           VARCHAR(255) DEFAULT NULL,
    `pemilik_proyek`      VARCHAR(255) DEFAULT NULL,
    `lokasi`              VARCHAR(500) DEFAULT NULL,
    `no_kontrak`          VARCHAR(100) DEFAULT NULL,
    `tgl_mulai`           DATE DEFAULT NULL,
    `pagu_kontrak_total`  DECIMAL(20,2) DEFAULT 0.00,
    `alamat_kontraktor`   TEXT DEFAULT NULL,
    `telp_kontraktor`     VARCHAR(50) DEFAULT NULL,
    `logo_perusahaan`     VARCHAR(255) DEFAULT NULL,
    `logo_pemilik`        VARCHAR(255) DEFAULT NULL,
    `created_at`          DATETIME DEFAULT CURRENT_TIMESTAMP
) ENGINE=InnoDB;

CREATE TABLE IF NOT EXISTS `laporan_harian` (
    `id`              INT AUTO_INCREMENT PRIMARY KEY,
    `proyek_id`       INT NOT NULL,
    `tanggal_laporan` DATE NOT NULL,
    `cuaca`           VARCHAR(100) DEFAULT NULL,
    `dokumen_upload`  VARCHAR(255) DEFAULT NULL,
    `created_at`      DATETIME DEFAULT CURRENT_TIMESTAMP,
    FOREIGN KEY (`proyek_id`) REFERENCES `master_proyek`(`id`) ON DELETE CASCADE
) ENGINE=InnoDB;

CREATE TABLE IF NOT EXISTS `tenaga_kerja` (
    `id`             INT AUTO_INCREMENT PRIMARY KEY,
    `laporan_id`     INT NOT NULL,
    `jenis_pekerja`  VARCHAR(150) DEFAULT NULL,
    `jumlah`         INT DEFAULT 0,
    `hadir`          INT DEFAULT 0,
    `tidak_hadir`    INT DEFAULT 0,
    `keterangan`     VARCHAR(255) DEFAULT NULL,
    FOREIGN KEY (`laporan_id`) REFERENCES `laporan_harian`(`id`) ON DELETE CASCADE
) ENGINE=InnoDB;

CREATE TABLE IF NOT EXISTS `peralatan` (
    `id`            INT AUTO_INCREMENT PRIMARY KEY,
    `laporan_id`    INT NOT NULL,
    `nama_alat`     VARCHAR(200) DEFAULT NULL,
    `jumlah_alat`   INT DEFAULT 0,
    `kondisi`       VARCHAR(100) DEFAULT NULL,
    `keterangan`    VARCHAR(255) DEFAULT NULL,
    FOREIGN KEY (`laporan_id`) REFERENCES `laporan_harian`(`id`) ON DELETE CASCADE
) ENGINE=InnoDB;

CREATE TABLE IF NOT EXISTS `material` (
    `id`               INT AUTO_INCREMENT PRIMARY KEY,
    `laporan_id`       INT NOT NULL,
    `nama_material`    VARCHAR(200) DEFAULT NULL,
    `volume_datang`    DECIMAL(12,2) DEFAULT 0,
    `satuan`           VARCHAR(50) DEFAULT NULL,
    `keterangan`       VARCHAR(255) DEFAULT NULL,
    FOREIGN KEY (`laporan_id`) REFERENCES `laporan_harian`(`id`) ON DELETE CASCADE
) ENGINE=InnoDB;

CREATE TABLE IF NOT EXISTS `pekerjaan` (
    `id`                 INT AUTO_INCREMENT PRIMARY KEY,
    `laporan_id`         INT NOT NULL,
    `jenis_pekerjaan`    VARCHAR(255) DEFAULT NULL,
    `lokasi`             VARCHAR(255) DEFAULT NULL,
    `vol_harian`         VARCHAR(100) DEFAULT NULL,
    `proses_kumulatif`   VARCHAR(50) DEFAULT NULL,
    `target_harian`      VARCHAR(50) DEFAULT NULL,
    `keterangan`         VARCHAR(255) DEFAULT NULL,
    FOREIGN KEY (`laporan_id`) REFERENCES `laporan_harian`(`id`) ON DELETE CASCADE
) ENGINE=InnoDB;

CREATE TABLE IF NOT EXISTS `kondisi_lapangan` (
    `id`                 INT AUTO_INCREMENT PRIMARY KEY,
    `laporan_id`         INT NOT NULL UNIQUE,
    `akses`              VARCHAR(255) DEFAULT NULL,
    `k3`                 VARCHAR(255) DEFAULT NULL,
    `kondisi_fisik`      VARCHAR(255) DEFAULT NULL,
    `hambatan`           VARCHAR(255) DEFAULT NULL,
    `tak_terencana`      TEXT DEFAULT NULL,
    FOREIGN KEY (`laporan_id`) REFERENCES `laporan_harian`(`id`) ON DELETE CASCADE
) ENGINE=InnoDB;

CREATE TABLE IF NOT EXISTS `pengesahan` (
    `id`               INT AUTO_INCREMENT PRIMARY KEY,
    `laporan_id`       INT NOT NULL UNIQUE,
    `nama_pembuat`     VARCHAR(200) DEFAULT NULL,
    `nama_penyetuju`   VARCHAR(200) DEFAULT NULL,
    FOREIGN KEY (`laporan_id`) REFERENCES `laporan_harian`(`id`) ON DELETE CASCADE
) ENGINE=InnoDB;

CREATE TABLE IF NOT EXISTS `master_wbs` (
    `id`              INT AUTO_INCREMENT PRIMARY KEY,
    `proyek_id`       INT NOT NULL,
    `kode_wbs`        VARCHAR(50) NOT NULL,
    `nama_pekerjaan`  VARCHAR(255) NOT NULL,
    `bobot_persen`    DECIMAL(6,2) DEFAULT 0.00,
    FOREIGN KEY (`proyek_id`) REFERENCES `master_proyek`(`id`) ON DELETE CASCADE
) ENGINE=InnoDB;

CREATE TABLE IF NOT EXISTS `kategori_budget` (
    `id`             INT AUTO_INCREMENT PRIMARY KEY,
    `proyek_id`      INT DEFAULT NULL,
    `nama_kategori`  VARCHAR(100) NOT NULL,
    `anggaran_pagu`  DECIMAL(20,2) DEFAULT 0.00,
    UNIQUE KEY `uniq_kategori_per_proyek` (`proyek_id`, `nama_kategori`),
    FOREIGN KEY (`proyek_id`) REFERENCES `master_proyek`(`id`) ON DELETE CASCADE
) ENGINE=InnoDB;

CREATE TABLE IF NOT EXISTS `master_budget` (
    `id`                 INT AUTO_INCREMENT PRIMARY KEY,
    `proyek_id`          INT NOT NULL,
    `kategori`           VARCHAR(100) DEFAULT NULL,
    `nama_item`          VARCHAR(255) DEFAULT NULL,
    `anggaran_pagu`      DECIMAL(20,2) DEFAULT 0.00,
    `realisasi_biaya`    DECIMAL(20,2) DEFAULT 0.00,
    `no_transaksi`       VARCHAR(50) DEFAULT NULL,
    `tanggal_transaksi`  DATE DEFAULT NULL,
    `jenis_kas`          ENUM('masuk','keluar') DEFAULT 'keluar',
    `pembuat`            VARCHAR(100) DEFAULT NULL,
    `kuantitas`          DECIMAL(12,2) DEFAULT 1.00,
    `harga_satuan`       DECIMAL(20,2) DEFAULT 0.00,
    `created_at`         DATETIME DEFAULT CURRENT_TIMESTAMP,
    FOREIGN KEY (`proyek_id`) REFERENCES `master_proyek`(`id`) ON DELETE CASCADE
) ENGINE=InnoDB;

CREATE TABLE IF NOT EXISTS `menu_permissions` (
    `id`         INT AUTO_INCREMENT PRIMARY KEY,
    `role`       ENUM('admin','manager','user') NOT NULL,
    `menu_key`   VARCHAR(50) NOT NULL,
    `can_access` TINYINT(1) NOT NULL DEFAULT 1,
    UNIQUE KEY `uniq_role_menu` (`role`, `menu_key`)
) ENGINE=InnoDB;

CREATE TABLE IF NOT EXISTS `active_sessions` (
    `id`            INT AUTO_INCREMENT PRIMARY KEY,
    `user_id`       INT NOT NULL,
    `session_token` VARCHAR(255) NOT NULL,
    `ip_address`    VARCHAR(50) DEFAULT NULL,
    `user_agent`    TEXT DEFAULT NULL,
    `device_type`   VARCHAR(50) DEFAULT NULL,
    `browser`       VARCHAR(100) DEFAULT NULL,
    `os`            VARCHAR(100) DEFAULT NULL,
    `login_at`      DATETIME DEFAULT CURRENT_TIMESTAMP,
    `last_active`   DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
    `is_active`     TINYINT(1) NOT NULL DEFAULT 1,
    KEY `idx_user_id` (`user_id`),
    KEY `idx_session_token` (`session_token`),
    KEY `idx_is_active` (`is_active`)
) ENGINE=InnoDB;

CREATE TABLE IF NOT EXISTS `user_projects` (
    `id`         INT AUTO_INCREMENT PRIMARY KEY,
    `user_id`    INT NOT NULL,
    `proyek_id`  INT NOT NULL,
    UNIQUE KEY `uniq_user_proyek` (`user_id`, `proyek_id`),
    FOREIGN KEY (`user_id`) REFERENCES `users`(`id`) ON DELETE CASCADE,
    FOREIGN KEY (`proyek_id`) REFERENCES `master_proyek`(`id`) ON DELETE CASCADE
) ENGINE=InnoDB;

CREATE TABLE IF NOT EXISTS `user_permissions` (
    `id`              INT AUTO_INCREMENT PRIMARY KEY,
    `user_id`         INT NOT NULL UNIQUE,
    `can_edit`        TINYINT(1) NOT NULL DEFAULT 1,
    `can_delete`      TINYINT(1) NOT NULL DEFAULT 1,
    `can_add_project` TINYINT(1) NOT NULL DEFAULT 0,
    FOREIGN KEY (`user_id`) REFERENCES `users`(`id`) ON DELETE CASCADE
) ENGINE=InnoDB;
"""

    try:
        # Pertama, koneksi tanpa database untuk buat database jika belum ada
        import urllib.parse
        db_host = os.environ.get('MYSQLHOST', '') or os.environ.get('DB_HOST', '127.0.0.1')
        db_port = int(os.environ.get('MYSQLPORT', '') or os.environ.get('DB_PORT', '3306') or '3306')
        db_user = os.environ.get('MYSQLUSER', '') or os.environ.get('DB_USER', 'root')
        db_password = os.environ.get('MYSQLPASSWORD', '') or os.environ.get('DB_PASSWORD', '')
        db_name = os.environ.get('MYSQLDATABASE', '') or os.environ.get('DB_NAME', 'db_proyek')
        
        # Parse dari DATABASE_URL jika ada
        database_url = os.environ.get('DATABASE_URL', '') or os.environ.get('MYSQL_URL', '') or os.environ.get('MYSQLURL', '')
        if database_url:
            if database_url.startswith('postgres://'):
                database_url = database_url.replace('postgres://', 'mysql://', 1)
            if database_url.startswith('mysql://'):
                parsed = urllib.parse.urlparse(database_url)
                db_host = parsed.hostname or db_host
                db_port = parsed.port or db_port
                db_user = parsed.username or db_user
                db_password = parsed.password or db_password
                db_name = (parsed.path or '').lstrip('/') or db_name
                if db_password:
                    db_password = urllib.parse.unquote(db_password)
        
        print(f'[MIGRATE] Connecting to: {db_host}:{db_port}/{db_name} as {db_user} (pwd_len={len(db_password)})')
        
        # Koneksi tanpa database dulu (SSL config untuk TiDB Cloud)
        conn_no_db = mysql.connector.connect(
            host=db_host, port=db_port, user=db_user, password=db_password,
            ssl_disabled=False, ssl_verify_cert=False, ssl_verify_identity=False
        )
        cursor_no_db = conn_no_db.cursor()
        
        # Buat database jika belum ada
        cursor_no_db.execute(f'CREATE DATABASE IF NOT EXISTS `{db_name}`')
        conn_no_db.commit()
        print(f'[MIGRATE] Database `{db_name}` siap!')
        cursor_no_db.close()
        conn_no_db.close()
        
        # Sekarang koneksi ke database yang benar
        conn = get_db_connection()
        cursor = conn.cursor()

        # Jalankan semua CREATE TABLE IF NOT EXISTS
        for statement in SQL_SCHEMA.split(';'):
            stmt = statement.strip()
            if stmt:
                cursor.execute(stmt)
        conn.commit()
        print('[MIGRATE] Semua tabel berhasil dibuat/dicek!')

        # Seed default permissions (hanya jika tabel kosong)
        cursor.execute("SELECT COUNT(*) FROM menu_permissions")
        if cursor.fetchone()[0] == 0:
            all_menus = ['dashboard','pengaturan','users','active_sessions',
                         'daily_report','weekly_report','monthly_report','gantt','wbs','budget','sync']
            defaults = {
                'admin':   {m: 1 for m in all_menus},
                'manager': {m: (0 if m in ('users', 'active_sessions') else 1) for m in all_menus},
                'user':    {m: (1 if m in ('dashboard', 'daily_report', 'weekly_report', 'monthly_report', 'gantt') else 0) for m in all_menus}
            }
            for role, perms in defaults.items():
                for menu_key, can_access in perms.items():
                    cursor.execute(
                        "INSERT INTO menu_permissions (role, menu_key, can_access) VALUES (%s, %s, %s)",
                        (role, menu_key, can_access)
                    )
            conn.commit()
            print('[MIGRATE] Default permissions di-seed!')
        else:
            print('[MIGRATE] Permissions sudah ada, skip seed.')

        # Seed admin default (jika belum ada user)
        cursor.execute("SELECT COUNT(*) FROM users")
        if cursor.fetchone()[0] == 0:
            from werkzeug.security import generate_password_hash
            cursor.execute(
                "INSERT INTO users (nama_lengkap, username, password_hash, role, status) VALUES (%s, %s, %s, 'admin', 'approved')",
                ('Administrator', 'admin', generate_password_hash('admin123'))
            )
            conn.commit()
            print('[MIGRATE] Admin default dibuat (admin/admin123)')

        cursor.close()
        conn.close()
        print('[MIGRATE] Semua migrasi berhasil!')
    except Exception as e:
        print(f'[AUTO-MIGRATE] ERROR: {e}')

auto_migrate()

# ====================
# 0. SPLASH SCREEN & AUTENTIKASI
# ====================
@app.route('/')
def index():
    if session.get('logged_in'):
        return redirect(url_for('dashboard'))
    return redirect(url_for('splash'))

@app.route('/splash')
def splash():
    if session.get('logged_in'):
        return redirect(url_for('dashboard'))
    return render_template('splash.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('logged_in'):
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        if not validate_csrf(): return redirect(url_for('login'))
        username, password = request.form.get('username', '').strip(), request.form.get('password', '')
        if not username or not password:
            flash('Username dan password harus diisi!', 'warning')
            return render_template('login.html')
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
        user = cursor.fetchone()
        conn.close()
        
        if user and check_password_hash(user['password_hash'], password):
            if user.get('status', 'approved') == 'pending':
                flash('Akun menunggu persetujuan admin.', 'warning'); return render_template('login.html')
            elif user.get('status', 'approved') == 'rejected':
                flash('Akun ditolak.', 'danger'); return render_template('login.html')
            
            user_role = user.get('role', 'user')
            
            # ── Load Edit/Delete Permission ──
            can_edit, can_delete = True, True
            if user_role in ('user', 'manager'):
                try:
                    conn_p = get_db_connection()
                    cur_p = conn_p.cursor(dictionary=True)
                    cur_p.execute('SELECT can_edit, can_delete FROM user_permissions WHERE user_id = %s', (user['id'],))
                    perm = cur_p.fetchone()
                    cur_p.close(); conn_p.close()
                    if perm:
                        can_edit = bool(perm['can_edit'])
                        can_delete = bool(perm['can_delete'])
                except:
                    pass
            
            session.update({
                'logged_in': True, 
                'user_id': user['id'],
                'username': user['username'],
                'nama_lengkap': user.get('nama_lengkap', user['username']),
                'role': user_role,
                'can_edit': can_edit,
                'can_delete': can_delete,
                'menu_permissions': load_permissions(user_role)
            })
            
            # ── Catat Sesi Aktif ──
            try:
                ua_raw = request.headers.get('User-Agent', '')
                ua_parsed = parse_user_agent(ua_raw)
                client_ip = request.headers.get('X-Forwarded-For', request.remote_addr or '0.0.0.0')
                if ',' in client_ip:
                    client_ip = client_ip.split(',')[0].strip()
                sess_token = secrets.token_hex(32)
                session['session_token'] = sess_token
                
                conn_s = get_db_connection()
                cur_s = conn_s.cursor()
                cur_s.execute(
                    "INSERT INTO active_sessions (user_id, session_token, ip_address, user_agent, device_type, browser, os) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    (user['id'], sess_token, client_ip, ua_raw,
                     ua_parsed['device_type'], ua_parsed['browser'], ua_parsed['os'])
                )
                conn_s.commit()
                cur_s.close()
                conn_s.close()
            except Exception:
                pass  # Jangan gagalkan login hanya karena tracking error
            
            # Alihkan ke halaman Welcome Screen berdurasi 7 detik sebelum ke Dashboard
            return redirect(url_for('welcome_screen'))
        else:
            flash('Username atau password salah!', 'danger')
            return render_template('login.html')
    return render_template('login.html')

@app.route('/welcome')
@login_required
def welcome_screen():
    return render_template('welcome.html', nama=session.get('nama_lengkap', session.get('username', 'User')))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        if not validate_csrf(): return redirect(url_for('register'))
        nama, username, pwd = request.form.get('nama_lengkap', '').strip(), request.form.get('username', '').strip(), request.form.get('password', '')
        if not nama or not username or not pwd:
            flash('Semua field wajib diisi!', 'warning'); return render_template('register.html')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("INSERT INTO users (nama_lengkap, username, password_hash, role, status) VALUES (%s, %s, %s, 'user', 'pending')", 
                           (nama, username, generate_password_hash(pwd)))
            conn.commit()
            flash('Pendaftaran berhasil! Tunggu verifikasi admin.', 'success')
            return redirect(url_for('login'))
        except mysql.connector.IntegrityError:
            flash('Username sudah digunakan.', 'danger'); return render_template('register.html')
        finally:
            cursor.close(); conn.close()
    return render_template('register.html')

@app.route('/logout')
def logout():
    # Hapus sesi aktif dari DB sebelum clear session
    sess_token = session.get('session_token')
    if sess_token:
        try:
            conn_s = get_db_connection()
            cur_s = conn_s.cursor()
            cur_s.execute("UPDATE active_sessions SET is_active = 0 WHERE session_token = %s", (sess_token,))
            conn_s.commit()
            cur_s.close()
            conn_s.close()
        except Exception:
            pass
    session.clear()
    flash('Berhasil logout.', 'info')
    return redirect(url_for('login'))

# ==========================================
# 0.5 USER MANAGEMENT (Admin Only)
# ==========================================
@app.route('/users')
@role_required('admin')
def users_management():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users ORDER BY created_at DESC")
    users_list = cursor.fetchall()
    conn.close()
    return render_template('users.html', users_list=users_list)

@app.route('/users/approve/<int:user_id>', methods=['POST'])
@role_required('admin')
def approve_user(user_id):
    if not validate_csrf(): return redirect(url_for('users_management'))
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET status = 'approved' WHERE id = %s", (user_id,))
    conn.commit(); cursor.close(); conn.close()
    flash('User berhasil disetujui!', 'success')
    return redirect(url_for('users_management'))

@app.route('/users/reject/<int:user_id>', methods=['POST'])
@role_required('admin')
def reject_user(user_id):
    if not validate_csrf(): return redirect(url_for('users_management'))
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET status = 'rejected' WHERE id = %s", (user_id,))
    conn.commit(); cursor.close(); conn.close()
    flash('User ditolak.', 'warning')
    return redirect(url_for('users_management'))

@app.route('/users/delete/<int:user_id>', methods=['POST'])
@role_required('admin')
def delete_user(user_id):
    if not validate_csrf(): return redirect(url_for('users_management'))
    if session.get('user_id') == user_id:
        flash('Tidak bisa menghapus akun sendiri!', 'danger')
        return redirect(url_for('users_management'))
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
    conn.commit(); cursor.close(); conn.close()
    flash('User berhasil dihapus.', 'danger')
    return redirect(url_for('users_management'))

@app.route('/users/edit_role/<int:user_id>', methods=['POST'])
@role_required('admin')
def edit_user_role(user_id):
    if not validate_csrf(): return redirect(url_for('users_management'))
    new_role = request.form.get('new_role', 'user')
    if new_role not in ('admin', 'manager', 'user'):
        flash('Role tidak valid!', 'danger')
        return redirect(url_for('users_management'))
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET role = %s WHERE id = %s", (new_role, user_id))
    conn.commit(); cursor.close(); conn.close()
    flash(f'Role user berhasil diubah menjadi {new_role}.', 'success')
    return redirect(url_for('users_management'))

# ==========================================
# 0.5.1 USER PROJECT ACCESS & PERMISSIONS
# ==========================================

def can_user_access_project(user_id, proyek_id):
    """Cek apakah user boleh mengakses proyek tertentu. Admin selalu bisa."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Cek role user
        cursor.execute('SELECT role FROM users WHERE id = %s', (user_id,))
        user = cursor.fetchone()
        if not user or user['role'] == 'admin':
            cursor.close(); conn.close()
            return True
        
        # Cek apakah user punya record di user_projects
        cursor.execute('SELECT COUNT(*) as cnt FROM user_projects WHERE user_id = %s', (user_id,))
        total_records = cursor.fetchone()['cnt']
        
        if total_records == 0:
            # Belum ada pengaturan akses, izinkan semua proyek
            cursor.close(); conn.close()
            return True
        
        # Cek apakah user punya akses ke proyek ini
        cursor.execute('SELECT COUNT(*) as cnt FROM user_projects WHERE user_id = %s AND proyek_id = %s', (user_id, proyek_id))
        has_access = cursor.fetchone()['cnt'] > 0
        cursor.close(); conn.close()
        return has_access
    except:
        return True  # Jika tabel belum ada, izinkan semua

def can_user_edit(user_id):
    """Cek apakah user boleh edit data. Admin selalu bisa, manager/user cek permission."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute('SELECT role FROM users WHERE id = %s', (user_id,))
        user = cursor.fetchone()
        if not user or user['role'] == 'admin':
            cursor.close(); conn.close()
            return True
        
        cursor.execute('SELECT can_edit FROM user_permissions WHERE user_id = %s', (user_id,))
        perm = cursor.fetchone()
        cursor.close(); conn.close()
        # Default: manager bisa edit, user juga bisa (kecuali di-set read-only)
        if perm:
            return bool(perm['can_edit'])
        return True  # Belum ada record permission = boleh edit
    except:
        return True

def can_user_delete(user_id):
    """Cek apakah user boleh hapus data. Admin selalu bisa, manager/user cek permission."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute('SELECT role FROM users WHERE id = %s', (user_id,))
        user = cursor.fetchone()
        if not user or user['role'] == 'admin':
            cursor.close(); conn.close()
            return True
        
        cursor.execute('SELECT can_delete FROM user_permissions WHERE user_id = %s', (user_id,))
        perm = cursor.fetchone()
        cursor.close(); conn.close()
        if perm:
            return bool(perm['can_delete'])
        return True  # Belum ada record permission = boleh hapus
    except:
        return True

# Register helpers ke Jinja
app.jinja_env.globals['can_user_access_project'] = can_user_access_project
app.jinja_env.globals['can_user_edit'] = can_user_edit
app.jinja_env.globals['can_user_delete'] = can_user_delete

@app.route('/users/projects/<int:user_id>', methods=['GET', 'POST'])
@role_required('admin')
def manage_user_projects(user_id):
    """Atur akses proyek untuk user tertentu."""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Ambil info user
    cursor.execute('SELECT * FROM users WHERE id = %s', (user_id,))
    user = cursor.fetchone()
    if not user:
        flash('User tidak ditemukan!', 'danger')
        return redirect(url_for('users_management'))
    
    if request.method == 'POST':
        if not validate_csrf():
            return redirect(url_for('manage_user_projects', user_id=user_id))
        
        # Hapus semua akses lama
        cursor.execute('DELETE FROM user_projects WHERE user_id = %s', (user_id,))
        
        # Simpan akses baru
        proyek_ids = request.form.getlist('proyek_access[]')
        for pid in proyek_ids:
            cursor.execute('INSERT INTO user_projects (user_id, proyek_id) VALUES (%s, %s)', (user_id, int(pid)))
        
        # Simpan permission edit/delete
        can_edit = 1 if request.form.get('can_edit') else 0
        can_delete = 1 if request.form.get('can_delete') else 0
        can_add = 1 if request.form.get('can_add_project') else 0
        
        cursor.execute(
            'INSERT INTO user_permissions (user_id, can_edit, can_delete, can_add_project) VALUES (%s, %s, %s, %s) '
            'ON DUPLICATE KEY UPDATE can_edit=%s, can_delete=%s, can_add_project=%s',
            (user_id, can_edit, can_delete, can_add, can_edit, can_delete, can_add)
        )
        
        conn.commit(); cursor.close(); conn.close()
        flash(f'Hak akses proyek untuk {user["username"]} berhasil disimpan!', 'success')
        return redirect(url_for('users_management'))
    
    # Ambil semua proyek
    cursor.execute('SELECT * FROM master_proyek ORDER BY id DESC')
    all_proyeks = cursor.fetchall()
    
    # Ambil proyek yang sudah diakses user
    cursor.execute('SELECT proyek_id FROM user_projects WHERE user_id = %s', (user_id,))
    user_proyek_ids = [r['proyek_id'] for r in cursor.fetchall()]
    
    # Ambil permission user
    cursor.execute('SELECT * FROM user_permissions WHERE user_id = %s', (user_id,))
    user_perm = cursor.fetchone()
    
    cursor.close(); conn.close()
    
    return render_template('manage_user_projects.html',
                          user=user, all_proyeks=all_proyeks,
                          user_proyek_ids=user_proyek_ids, user_perm=user_perm)

@app.route('/users/readonly/<int:user_id>', methods=['POST'])
@role_required('admin')
def toggle_readonly(user_id):
    """Toggle mode read-only untuk user."""
    if not validate_csrf(): return redirect(url_for('users_management'))
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Ambil permission saat ini
    cursor.execute('SELECT * FROM user_permissions WHERE user_id = %s', (user_id,))
    perm = cursor.fetchone()
    
    new_edit = 0 if (perm and perm['can_edit']) else 1
    new_delete = 0 if (perm and perm['can_delete']) else 1
    
    # Ambil nama user
    cursor.execute('SELECT username FROM users WHERE id = %s', (user_id,))
    u = cursor.fetchone()
    uname = u['username'] if u else 'User'
    
    cursor.execute(
        'INSERT INTO user_permissions (user_id, can_edit, can_delete) VALUES (%s, %s, %s) '
        'ON DUPLICATE KEY UPDATE can_edit=%s, can_delete=%s',
        (user_id, new_edit, new_delete, new_edit, new_delete)
    )
    conn.commit(); cursor.close(); conn.close()
    
    status = 'READ-ONLY 👁️' if new_edit == 0 else 'FULL ACCESS ✏️'
    flash(f'User {uname} sekarang dalam mode: {status}', 'info')
    return redirect(url_for('users_management'))

# ==========================================
# 0.6 HAK AKSES / MENU PERMISSIONS (Admin Only)
# ==========================================
MENU_DEFINITIONS = [
    {'key': 'dashboard',        'label': 'Dashboard Utama',         'icon': 'fa-house',              'group': 'Utama'},
    {'key': 'pengaturan',       'label': 'Pengaturan Utama',        'icon': 'fa-gear',               'group': 'Utama'},
    {'key': 'users',            'label': 'Kelola User',             'icon': 'fa-users-gear',         'group': 'Utama'},
    {'key': 'active_sessions',  'label': 'Sesi Aktif',              'icon': 'fa-desktop',            'group': 'Utama'},
    {'key': 'daily_report',     'label': 'Daily Report',            'icon': 'fa-clock-rotate-left',  'group': 'Proyek'},
    {'key': 'weekly_report',    'label': 'Weekly Report',           'icon': 'fa-calendar-week',      'group': 'Proyek'},
    {'key': 'monthly_report',   'label': 'Monthly Report',          'icon': 'fa-calendar-days',      'group': 'Proyek'},
    {'key': 'gantt',            'label': 'Time Schedule & Kurva S', 'icon': 'fa-chart-line',          'group': 'Proyek'},
    {'key': 'wbs',              'label': 'WBS Management',          'icon': 'fa-sitemap',            'group': 'Proyek'},
    {'key': 'budget',           'label': 'Cost & Budget',           'icon': 'fa-sack-dollar',        'group': 'Proyek'},
    {'key': 'sync',             'label': 'Sync Data',               'icon': 'fa-arrows-rotate',      'group': 'Sistem'},
]

@app.route('/hak_akses')
@role_required('admin')
def hak_akses():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Ambil semua permission yang ada
    cursor.execute("SELECT role, menu_key, can_access FROM menu_permissions ORDER BY role, id")
    all_perms = cursor.fetchall()
    conn.close()
    
    # Susun ke dict: {role: {menu_key: bool}}
    perm_matrix = {'admin': {}, 'manager': {}, 'user': {}}
    for p in all_perms:
        perm_matrix[p['role']][p['menu_key']] = bool(p['can_access'])
    
    # Pastikan admin selalu full access
    for m in MENU_DEFINITIONS:
        perm_matrix['admin'][m['key']] = True
    
    return render_template('hak_akses.html', menu_definitions=MENU_DEFINITIONS, perm_matrix=perm_matrix)

@app.route('/hak_akses/save', methods=['POST'])
@role_required('admin')
def save_hak_akses():
    if not validate_csrf(): return redirect(url_for('hak_akses'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        roles = ['admin', 'manager', 'user']
        menu_keys = [m['key'] for m in MENU_DEFINITIONS]
        
        for role in roles:
            for menu_key in menu_keys:
                # Admin selalu full access
                can_access = 1 if role == 'admin' else (1 if request.form.get(f'perm_{role}_{menu_key}') else 0)
                
                cursor.execute(
                    "INSERT INTO menu_permissions (role, menu_key, can_access) VALUES (%s, %s, %s) "
                    "ON DUPLICATE KEY UPDATE can_access = %s",
                    (role, menu_key, can_access, can_access)
                )
        
        conn.commit()
        
        # Refresh permissions di session admin yang sedang login
        session['menu_permissions'] = load_permissions(session.get('role', 'admin'))
        
        flash('Hak Akses berhasil disimpan!', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Gagal menyimpan hak akses: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('hak_akses'))

# ==========================================
# 0.7 SESI AKTIF PENGGUNA (Admin Only)
# ==========================================
@app.route('/active_sessions')
@role_required('admin')
def active_sessions_view():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Ambil semua sesi aktif + info user
    sql = """
        SELECT 
            a.id, a.session_token, a.ip_address, a.user_agent,
            a.device_type, a.browser, a.os, a.login_at, a.last_active, a.is_active,
            u.id AS uid, u.nama_lengkap, u.username, u.role
        FROM active_sessions a
        JOIN users u ON a.user_id = u.id
        ORDER BY a.is_active DESC, a.last_active DESC
    """
    cursor.execute(sql)
    all_sessions = cursor.fetchall()
    
    # Hitung statistik
    active_count = sum(1 for s in all_sessions if s.get('is_active'))
    total_count = len(all_sessions)
    
    # Ambil jumlah user unik yang sedang aktif
    active_user_ids = set()
    for s in all_sessions:
        if s.get('is_active'):
            active_user_ids.add(s['uid'])
    
    cursor.execute("SELECT COUNT(*) as cnt FROM users WHERE status = 'approved'")
    total_users = cursor.fetchone()['cnt']
    
    conn.close()
    
    return render_template('active_sessions.html',
                           sessions=all_sessions,
                           active_count=active_count,
                           total_count=total_count,
                           online_users=len(active_user_ids),
                           total_users=total_users)

@app.route('/active_sessions/kick/<int:session_id>', methods=['POST'])
@role_required('admin')
def kick_session(session_id):
    """Force logout sesi tertentu."""
    if not validate_csrf(): return redirect(url_for('active_sessions_view'))
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE active_sessions SET is_active = 0 WHERE id = %s", (session_id,))
    conn.commit()
    cursor.close()
    conn.close()
    flash('Sesi pengguna berhasil dipaksa logout.', 'danger')
    return redirect(url_for('active_sessions_view'))

@app.route('/active_sessions/kick_all', methods=['POST'])
@role_required('admin')
def kick_all_sessions():
    """Force logout semua sesi aktif (kecuali admin sendiri)."""
    if not validate_csrf(): return redirect(url_for('active_sessions_view'))
    my_token = session.get('session_token', '')
    conn = get_db_connection()
    cursor = conn.cursor()
    if my_token:
        cursor.execute("UPDATE active_sessions SET is_active = 0 WHERE is_active = 1 AND session_token != %s", (my_token,))
    else:
        cursor.execute("UPDATE active_sessions SET is_active = 0 WHERE is_active = 1")
    affected = cursor.rowcount
    conn.commit()
    cursor.close()
    conn.close()
    flash(f'{affected} sesi pengguna lain berhasil dipaksa logout.', 'danger')
    return redirect(url_for('active_sessions_view'))

@app.route('/active_sessions/cleanup', methods=['POST'])
@role_required('admin')
def cleanup_sessions():
    """Hapus semua record sesi yang sudah tidak aktif."""
    if not validate_csrf(): return redirect(url_for('active_sessions_view'))
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM active_sessions WHERE is_active = 0")
    affected = cursor.rowcount
    conn.commit()
    cursor.close()
    conn.close()
    flash(f'{affected} record sesi lama berhasil dibersihkan.', 'info')
    return redirect(url_for('active_sessions_view'))

# ==========================================
# 0.8 SYNC DATA (Local ↔ Cloud)
# ==========================================

def get_cloud_connection():
    """Koneksi ke TiDB Cloud untuk sync."""
    import urllib.parse
    database_url = os.environ.get('DATABASE_URL', '') or os.environ.get('MYSQL_URL', '') or os.environ.get('MYSQLURL', '')
    
    # Default values
    host = os.environ.get('MYSQLHOST', '') or ''
    port = int(os.environ.get('MYSQLPORT', '4000') or '4000')
    user = os.environ.get('MYSQLUSER', '') or ''
    password = os.environ.get('MYSQLPASSWORD', '') or ''
    dbname = os.environ.get('MYSQLDATABASE', '') or 'db_proyek'
    
    if database_url:
        try:
            if database_url.startswith('postgres://'):
                database_url = database_url.replace('postgres://', 'mysql://', 1)
            if database_url.startswith('mysql://'):
                parsed = urllib.parse.urlparse(database_url)
                host = parsed.hostname or host
                port = parsed.port or port
                user = parsed.username or user
                password = urllib.parse.unquote(parsed.password or '') or password
                dbname = (parsed.path or '').lstrip('/') or dbname
        except:
            pass
    
    if host and user:
        try:
            conn = mysql.connector.connect(
                host=host, port=port, user=user, password=password,
                database=dbname, ssl_disabled=False,
                ssl_verify_cert=False, ssl_verify_identity=False
            )
            return conn
        except:
            return None
    return None

# Tabel yang bisa di-sync (dalam urutan yang benar karena foreign key)
SYNC_TABLES = [
    'users', 'master_proyek', 'master_kategori_biaya',
    'laporan_harian', 'tenaga_kerja', 'peralatan', 'material',
    'pekerjaan', 'kondisi_lapangan', 'pengesahan',
    'master_wbs', 'kategori_budget', 'master_budget',
    'menu_permissions', 'settings', 'user_projects',
    'user_permissions', 'kategori_biaya'
]

# Natural key per tabel — kolom unik untuk deteksi duplikat
# Digunakan untuk cek apakah record sudah ada sebelum INSERT
NATURAL_KEYS = {
    'users': 'username',
    'master_proyek': 'no_kontrak',
    'master_kategori_biaya': 'nama_kategori',
    'laporan_harian': None,  # pakai (proyek_id, tanggal_laporan) combo
    'tenaga_kerja': None,    # pakai laporan_id + jenis_pekerja combo
    'peralatan': None,
    'material': None,
    'pekerjaan': None,
    'kondisi_lapangan': None,
    'pengesahan': None,
    'master_wbs': None,       # pakai (proyek_id, kode_wbs)
    'kategori_budget': None,  # pakai (proyek_id, nama_kategori)
    'master_budget': None,    # pakai (proyek_id, no_transaksi) if exists
    'menu_permissions': None, # pakai (role, menu_key)
    'settings': 'setting_key',
    'user_projects': None,    # pakai (user_id, proyek_id)
    'user_permissions': 'user_id',
    'kategori_biaya': None,
}

# Composite keys — kombinasi kolom yang jadi "id unik"
COMPOSITE_KEYS = {
    'laporan_harian': ['proyek_id', 'tanggal_laporan'],
    'tenaga_kerja': ['laporan_id', 'jenis_pekerja'],
    'peralatan': ['laporan_id', 'nama_peralatan'],
    'material': ['laporan_id', 'nama_material'],
    'pekerjaan': ['laporan_id', 'jenis_pekerjaan', 'lokasi_pekerjaan'],
    'kondisi_lapangan': ['laporan_id'],
    'pengesahan': ['laporan_id'],
    'master_wbs': ['proyek_id', 'kode_wbs'],
    'kategori_budget': ['proyek_id', 'nama_kategori'],
    'master_budget': ['proyek_id', 'no_transaksi'],
    'menu_permissions': ['role', 'menu_key'],
    'user_projects': ['user_id', 'proyek_id'],
    'kategori_biaya': ['proyek_id', 'nama_kategori'],
}

@app.route('/sync')
@role_required('admin')
def sync_data():
    """Halaman Sinkronisasi Data."""
    # Cek status koneksi cloud
    cloud_ok = False
    cloud_info = ''
    try:
        conn_cloud = get_cloud_connection()
        if conn_cloud:
            cur = conn_cloud.cursor()
            cur.execute('SELECT 1')
            cur.fetchone()
            cur.close()
            conn_cloud.close()
            cloud_ok = True
            cloud_info = 'Terhubung ke TiDB Cloud'
        else:
            cloud_info = 'DATABASE_URL belum di-set'
    except Exception as e:
        cloud_info = f'Error: {str(e)[:50]}'
    
    # Hitung jumlah data di lokal
    local_counts = {}
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        for table in SYNC_TABLES:
            try:
                cur.execute(f'SELECT COUNT(*) FROM `{table}`')
                local_counts[table] = cur.fetchone()[0]
            except:
                local_counts[table] = 0
        cur.close()
        conn.close()
    except:
        pass
    
    # Hitung jumlah data di cloud
    cloud_counts = {}
    if cloud_ok:
        try:
            conn_cloud = get_cloud_connection()
            cur = conn_cloud.cursor()
            for table in SYNC_TABLES:
                try:
                    cur.execute(f'SELECT COUNT(*) FROM `{table}`')
                    cloud_counts[table] = cur.fetchone()[0]
                except:
                    cloud_counts[table] = 0
            cur.close()
            conn_cloud.close()
        except:
            pass
    
    # Di cloud, auto-sync tidak aktif (tidak ada MySQL lokal)
    if is_cloud_environment():
        _sync_state['auto_sync_enabled'] = False
    
    return render_template('sync_data.html',
                          cloud_ok=cloud_ok, cloud_info=cloud_info,
                          local_counts=local_counts, cloud_counts=cloud_counts,
                          sync_tables=SYNC_TABLES)

@app.route('/sync/push', methods=['POST'])
@role_required('admin')
def sync_push():
    """Push data dari Lokal → Cloud."""
    if not validate_csrf(): return redirect(url_for('sync_data'))
    if is_cloud_environment():
        flash('Sync hanya bisa dijalankan dari komputer lokal!', 'warning')
        return redirect(url_for('sync_data'))
    
    conn_cloud = get_cloud_connection()
    if not conn_cloud:
        flash('Tidak bisa koneksi ke cloud! Pastikan DATABASE_URL sudah benar.', 'danger')
        return redirect(url_for('sync_data'))
    
    try:
        conn_local = get_db_connection()
        total = 0
        for table in SYNC_TABLES:
            try:
                n = _sync_table_push(conn_local, conn_cloud, table)
                total += n
                if n > 0:
                    print(f'[SYNC PUSH] {table}: {n} rows')
            except Exception as e:
                print(f'[SYNC PUSH] {table}: {e}')
        
        conn_local.close()
        conn_cloud.close()
        
        flash(f'Push berhasil! {total} data dikirim ke TiDB Cloud (safe merge).', 'success')
    except Exception as e:
        flash(f'Gagal push data: {str(e)}', 'danger')
    
    return redirect(url_for('sync_data'))

@app.route('/sync/pull', methods=['POST'])
@role_required('admin')
def sync_pull():
    """Pull data dari Cloud → Lokal."""
    if not validate_csrf(): return redirect(url_for('sync_data'))
    if is_cloud_environment():
        flash('Sync hanya bisa dijalankan dari komputer lokal!', 'warning')
        return redirect(url_for('sync_data'))
    
    conn_cloud = get_cloud_connection()
    if not conn_cloud:
        flash('Tidak bisa koneksi ke cloud! Pastikan DATABASE_URL sudah benar.', 'danger')
        return redirect(url_for('sync_data'))
    
    try:
        conn_local = get_db_connection()
        total = 0
        for table in SYNC_TABLES:
            try:
                n = _sync_table_pull(conn_local, conn_cloud, table)
                total += n
                if n > 0:
                    print(f'[SYNC PULL] {table}: {n} rows')
            except Exception as e:
                print(f'[SYNC PULL] {table}: {e}')
        
        conn_local.close()
        conn_cloud.close()
        
        flash(f'Pull berhasil! {total} data diambil dari TiDB Cloud (safe merge).', 'success')
    except Exception as e:
        flash(f'Gagal pull data: {str(e)}', 'danger')
    
    return redirect(url_for('sync_data'))

@app.route('/sync/bidirectional', methods=['POST'])
@role_required('admin')
def sync_bidirectional():
    """Sync bidirectional - gabungkan data dari lokal dan cloud."""
    if not validate_csrf(): return redirect(url_for('sync_data'))
    if is_cloud_environment():
        flash('Sync hanya bisa dijalankan dari komputer lokal!', 'warning')
        return redirect(url_for('sync_data'))
    
    conn_cloud = get_cloud_connection()
    if not conn_cloud:
        flash('Tidak bisa koneksi ke cloud! Pastikan DATABASE_URL sudah benar.', 'danger')
        return redirect(url_for('sync_data'))
    
    try:
        conn_local = get_db_connection()
        total = 0
        for table in SYNC_TABLES:
            try:
                # Push lokal -> cloud
                n1 = _sync_table_push(conn_local, conn_cloud, table)
                total += n1
                # Pull cloud -> lokal
                n2 = _sync_table_pull(conn_local, conn_cloud, table)
                total += n2
            except Exception as e:
                print(f'[SYNC BI] {table}: {e}')
        
        conn_local.close()
        conn_cloud.close()
        
        flash(f'Sync bidirectional berhasil! {total} data digabungkan (safe merge).', 'success')
    except Exception as e:
        flash(f'Gagal sync: {str(e)}', 'danger')
    
    return redirect(url_for('sync_data'))

# ====================
# DASHBOARD DLL
# ====================
@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    user_role = session.get('role', 'user')
    user_id = session.get('user_id')
    
    if user_role in ('admin', 'manager'):
        # Admin/manager lihat semua proyek
        cursor.execute("SELECT * FROM master_proyek ORDER BY id DESC")
    else:
        # User biasa: cek apakah punya user_projects record
        try:
            cursor.execute('SELECT COUNT(*) as cnt FROM user_projects WHERE user_id = %s', (user_id,))
            has_records = cursor.fetchone()['cnt'] > 0
            
            if has_records:
                # Hanya lihat proyek yang diakses
                cursor.execute("""
                    SELECT DISTINCT mp.* FROM master_proyek mp
                    JOIN user_projects up ON mp.id = up.proyek_id
                    WHERE up.user_id = %s
                    ORDER BY mp.id DESC
                """, (user_id,))
            else:
                # Belum ada pengaturan akses, tampilkan semua proyek
                cursor.execute("SELECT * FROM master_proyek ORDER BY id DESC")
        except:
            # Jika tabel user_projects belum ada, tampilkan semua
            cursor.execute("SELECT * FROM master_proyek ORDER BY id DESC")
    
    proyek_list = cursor.fetchall()
    conn.close()
    return render_template('dashboard.html', proyek_list=proyek_list)

@app.route('/add_project', methods=['POST'])
@role_required('admin', 'manager')
def add_project():
    if not validate_csrf(): return redirect(url_for('dashboard'))
    # ── Permission Check ──
    if not can_user_edit(session['user_id']):
        flash('Anda tidak memiliki akses untuk menambah proyek!', 'danger')
        return redirect(url_for('dashboard'))
    conn = get_db_connection()
    cursor = conn.cursor()
    sql = """INSERT INTO master_proyek (nama_kegiatan, nama_proyek, penyedia_jasa, konsultan, pemilik_proyek, lokasi, no_kontrak, tgl_mulai, pagu_kontrak_total) 
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"""
    val = (request.form['nama_kegiatan'], request.form['nama_proyek'], request.form['penyedia_jasa'], request.form['konsultan'], request.form['pemilik_proyek'], request.form['lokasi'], request.form['no_kontrak'], request.form['tgl_mulai'], request.form.get('pagu_kontrak_total', 0))
    cursor.execute(sql, val)
    proyek_id = cursor.lastrowid
    
    default_kategori = ['STAFF', 'UPAH MP', 'MATERIAL', 'MCU', 'PERALATAN', 'TRANSPORT & AKOMODASI', 'JASA SUBCONT', 'CONSUMABLE', 'ENTERTAIN', 'LAIN-LAIN']
    for kat in default_kategori:
        cursor.execute("INSERT INTO kategori_budget (proyek_id, nama_kategori, anggaran_pagu) VALUES (%s, %s, 0)", (proyek_id, kat))
        
    conn.commit()
    cursor.close()
    conn.close()
    flash('Proyek Baru Berhasil Ditambahkan dengan Kategori Biaya Default!', 'success')
    return redirect(url_for('dashboard'))

# ==========================================
# 1.5 PENGATURAN UTAMA & PROYEK
# ==========================================
@app.route('/pengaturan', methods=['GET', 'POST'])
@role_required('admin', 'manager')
def pengaturan_global():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM master_proyek ORDER BY id ASC LIMIT 1")
    proyek = cursor.fetchone()
    
    if request.method == 'POST':
        if not validate_csrf():
            conn.close()
            return redirect(url_for('pengaturan_global'))
        
        penyedia_jasa = request.form.get('penyedia_jasa')
        alamat_kontraktor = request.form.get('alamat_kontraktor')
        telp_kontraktor = request.form.get('telp_kontraktor')
        
        file = request.files.get('logo')
        if proyek:
            if file and file.filename != '':
                filename = secure_filename(f"logo_kontraktor_{file.filename}")
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                
                cursor.execute("""UPDATE master_proyek
                                  SET penyedia_jasa = %s, alamat_kontraktor = %s, telp_kontraktor = %s, logo_perusahaan = %s
                                  WHERE id = %s""",
                               (penyedia_jasa, alamat_kontraktor, telp_kontraktor, filename, proyek['id']))
            else:
                cursor.execute("""UPDATE master_proyek
                                  SET penyedia_jasa = %s, alamat_kontraktor = %s, telp_kontraktor = %s
                                  WHERE id = %s""",
                               (penyedia_jasa, alamat_kontraktor, telp_kontraktor, proyek['id']))
            conn.commit()
            flash('Pengaturan Utama Berhasil Diperbarui!', 'success')
        conn.close()
        return redirect(url_for('pengaturan_global'))
        
    conn.close()
    return render_template('pengaturan_utama.html', proyek=proyek)

@app.route('/pengaturan_proyek/<int:proyek_id>', methods=['GET', 'POST'])
@role_required('admin', 'manager')
def pengaturan_proyek(proyek_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    proyek = cursor.fetchone()
    
    cursor.execute("SELECT * FROM kategori_budget WHERE proyek_id = %s ORDER BY id ASC", (proyek_id,))
    kategori_list = cursor.fetchall()
    
    if request.method == 'POST':
        if not validate_csrf():
            conn.close()
            return redirect(url_for('pengaturan_proyek', proyek_id=proyek_id))
            
        nama_proyek = request.form.get('nama_proyek')
        pemilik_proyek = request.form.get('pemilik_proyek')
        lokasi = request.form.get('lokasi')
        no_kontrak = request.form.get('no_kontrak')
        
        file = request.files.get('logo_pemilik')
        if file and file.filename != '':
            filename_owner = secure_filename(f"logo_owner_{proyek_id}_{file.filename}")
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename_owner))
            
            cursor.execute("""UPDATE master_proyek
                              SET nama_proyek = %s, pemilik_proyek = %s, lokasi = %s, no_kontrak = %s, logo_pemilik = %s
                              WHERE id = %s""",
                           (nama_proyek, pemilik_proyek, lokasi, no_kontrak, filename_owner, proyek_id))
        else:
            cursor.execute("""UPDATE master_proyek
                              SET nama_proyek = %s, pemilik_proyek = %s, lokasi = %s, no_kontrak = %s
                              WHERE id = %s""",
                           (nama_proyek, pemilik_proyek, lokasi, no_kontrak, proyek_id))
            
        conn.commit()
        flash('Pengaturan Proyek Berhasil Diperbarui!', 'success')
        conn.close()
        return redirect(url_for('pengaturan_proyek', proyek_id=proyek_id))
        
    conn.close()
    return render_template('pengaturan_proyek.html', proyek=proyek, kategori_list=kategori_list)


# ==========================================
# 2. WORKSPACE & INPUT DAILY REPORT
# ==========================================
@app.route('/workspace/<int:proyek_id>')
@login_required
def workspace_history(proyek_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    proyek = cursor.fetchone()
    
    # Ambil filter params
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    jenis_filter = request.args.get('jenis_pekerjaan', '')
    
    # Build query - satu query dengan GROUP_CONCAT untuk jenis pekerjaan
    sql = """
        SELECT lh.*, 
               COALESCE(GROUP_CONCAT(DISTINCT p.jenis_pekerjaan SEPARATOR ', '), '-') as jenis_pekerjaan
        FROM laporan_harian lh
        LEFT JOIN pekerjaan p ON lh.id = p.laporan_id
        WHERE lh.proyek_id = %s
    """
    params = [proyek_id]
    
    if date_from:
        sql += " AND lh.tanggal_laporan >= %s"
        params.append(date_from)
    if date_to:
        sql += " AND lh.tanggal_laporan <= %s"
        params.append(date_to)
    if jenis_filter:
        sql += " AND p.jenis_pekerjaan = %s"
        params.append(jenis_filter)
    
    sql += " GROUP BY lh.id ORDER BY lh.tanggal_laporan DESC"
    cursor.execute(sql, params)
    laporan = cursor.fetchall()
    
    # Ambil daftar jenis pekerjaan unik untuk dropdown filter
    cursor.execute("""
        SELECT DISTINCT p.jenis_pekerjaan
        FROM pekerjaan p JOIN laporan_harian lh ON p.laporan_id = lh.id
        WHERE lh.proyek_id = %s ORDER BY p.jenis_pekerjaan
    """, (proyek_id,))
    jenis_list = [r['jenis_pekerjaan'] for r in cursor.fetchall()]
    
    conn.close()
    return render_template('daftar_laporan.html', proyek=proyek, laporan=laporan,
                           jenis_list=jenis_list, date_from=date_from, date_to=date_to,
                           jenis_filter=jenis_filter)

def _hitung_kumulatif_bulanan(cursor, proyek_id):
    """Hitung M1/M2/M3 dari kumulatif aktual laporan terakhir per bulan per jenis pekerjaan."""
    # Ambil kumulatif terakhir per bulan per jenis pekerjaan
    cursor.execute("""
        SELECT p.jenis_pekerjaan,
               DATE_FORMAT(lh.tanggal_laporan, '%%Y-%%m') as bulan,
               MAX(CAST(REPLACE(REPLACE(p.proses_kumulatif, '%%', ''), ',', '.') AS DECIMAL(7,2))) as kumulatif
        FROM pekerjaan p
        JOIN laporan_harian lh ON p.laporan_id = lh.id
        WHERE lh.proyek_id = %s AND p.proses_kumulatif IS NOT NULL AND p.proses_kumulatif != ''
        GROUP BY p.jenis_pekerjaan, bulan
        ORDER BY p.jenis_pekerjaan, bulan
    """, (proyek_id,))
    rows = cursor.fetchall()
    
    # Group by jenis pekerjaan
    data = {}  # {jenis: {bulan: kumulatif}}
    for r in rows:
        jp = r['jenis_pekerjaan']
        bln = r['bulan']
        kum = float(r['kumulatif']) if r['kumulatif'] else 0
        if jp not in data:
            data[jp] = {}
        data[jp][bln] = kum
    
    # Hitung M1/M2/M3 (kumulatif di akhir bulan ke-1, ke-2, ke-3 dari awal proyek)
    # Cari bulan awal proyek
    cursor.execute("SELECT MIN(tanggal_laporan) as mulai FROM laporan_harian WHERE proyek_id = %s", (proyek_id,))
    mulai = cursor.fetchone()
    if not mulai or not mulai['mulai']:
        return {}
    
    from datetime import datetime
    mulai_date = mulai['mulai']
    if isinstance(mulai_date, str):
        mulai_date = datetime.strptime(mulai_date, '%Y-%m-%d').date()
    
    # Generate bulan M1, M2, M3... dari bulan awal
    wbs_target_bulanan = {}
    for jp, bulan_data in data.items():
        targets = {}
        # M1 = bulan ke-1, M2 = bulan ke-2, M3 = bulan ke-3
        for m in range(1, 13):
            # Hitung bulan target
            target_month = mulai_date.month + m - 1
            target_year = mulai_date.year + (target_month - 1) // 12
            target_month = ((target_month - 1) % 12) + 1
            target_key = f'{target_year}-{target_month:02d}'
            
            # Cari kumulatif terakhir di bulan ini atau sebelumnya
            kum_val = 0
            for bln_key in sorted(bulan_data.keys()):
                if bln_key <= target_key:
                    kum_val = bulan_data[bln_key]
                else:
                    break
            if kum_val > 0:
                targets[f'M{m}'] = kum_val
        if targets:
            wbs_target_bulanan[jp] = targets
    
    return wbs_target_bulanan

@app.route('/input/<int:proyek_id>')
@login_required
def input_laporan(proyek_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    proyek = cursor.fetchone()
    # Ambil data WBS untuk dropdown jenis pekerjaan
    cursor.execute("SELECT nama_pekerjaan FROM master_wbs WHERE proyek_id = %s ORDER BY kode_wbs ASC", (proyek_id,))
    wbs_list = [r['nama_pekerjaan'] for r in cursor.fetchall()]
    
    # Hitung M1/M2/M3 dari data laporan aktual (kumulatif terakhir per bulan)
    wbs_target_bulanan = _hitung_kumulatif_bulanan(cursor, proyek_id)
    
    conn.close()
    g.wbs_list_cache = wbs_list
    return render_template('index.html', proyek=proyek, wbs_list=wbs_list,
                           wbs_target_bulanan=wbs_target_bulanan)

@app.route('/submit', methods=['POST'])
@login_required
def submit():
    if request.method == 'POST':
        if not validate_csrf(): return redirect(url_for('dashboard'))
        # ── Permission Check ──
        if not can_user_edit(session['user_id']):
            flash('Anda tidak memiliki akses untuk menambah data!', 'danger')
            return redirect(url_for('dashboard'))
        proyek_id = request.form['proyek_id']
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute("INSERT INTO laporan_harian (proyek_id, tanggal_laporan, cuaca) VALUES (%s, %s, %s)", 
                           (proyek_id, request.form['tanggal_laporan'], request.form['cuaca']))
            laporan_id = cursor.lastrowid
            
            jenis_pekerja = request.form.getlist('jenis_pekerja[]')
            for i in range(len(jenis_pekerja)):
                cursor.execute("INSERT INTO tenaga_kerja (laporan_id, jenis_pekerja, jumlah, hadir, tidak_hadir, keterangan) VALUES (%s, %s, %s, %s, %s, %s)",
                    (laporan_id, jenis_pekerja[i], request.form.getlist('jumlah_pekerja[]')[i], request.form.getlist('hadir[]')[i], request.form.getlist('tidak_hadir[]')[i], request.form.getlist('ket_pekerja[]')[i]))
                
            nama_alat = request.form.getlist('nama_alat[]')
            for i in range(len(nama_alat)):
                if nama_alat[i].strip() != '':
                    cursor.execute("INSERT INTO peralatan (laporan_id, nama_alat, jumlah_alat, kondisi, keterangan) VALUES (%s, %s, %s, %s, %s)",
                        (laporan_id, nama_alat[i], request.form.getlist('jumlah_alat[]')[i], request.form.getlist('kondisi_alat[]')[i], request.form.getlist('ket_alat[]')[i]))

            nama_material = request.form.getlist('nama_material[]')
            for i in range(len(nama_material)):
                if nama_material[i].strip() != '':
                    cursor.execute("INSERT INTO material (laporan_id, nama_material, volume_datang, satuan, keterangan) VALUES (%s, %s, %s, %s, %s)",
                        (laporan_id, nama_material[i], request.form.getlist('volume_material[]')[i], request.form.getlist('satuan_material[]')[i], request.form.getlist('ket_material[]')[i]))

            jenis_pekerjaan = request.form.getlist('jenis_pekerjaan[]')
            for i in range(len(jenis_pekerjaan)):
                cursor.execute("INSERT INTO pekerjaan (laporan_id, jenis_pekerjaan, lokasi, vol_harian, proses_kumulatif, target_harian, keterangan) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    (laporan_id, jenis_pekerjaan[i], request.form.getlist('lokasi_pekerjaan[]')[i], request.form.getlist('vol_harian[]')[i], request.form.getlist('proses_kumulatif[]')[i], request.form.getlist('target_harian[]')[i], request.form.getlist('ket_pekerjaan[]')[i]))
                
            cursor.execute("INSERT INTO kondisi_lapangan (laporan_id, akses, k3, kondisi_fisik, hambatan, tak_terencana) VALUES (%s, %s, %s, %s, %s, %s)",
                (laporan_id, request.form['akses_lapangan'], request.form['keselamatan'], request.form['kondisi_lapangan'], request.form['faktor_penghambat'], request.form['kegiatan_tak_terencana']))
            
            cursor.execute("INSERT INTO pengesahan (laporan_id, nama_pembuat, nama_penyetuju) VALUES (%s, %s, %s)",
                (laporan_id, request.form['nama_pembuat'], request.form['nama_penyetuju']))
            
            conn.commit()
            flash('Laporan Harian Berhasil Disimpan!', 'success')
            return redirect(url_for('workspace_history', proyek_id=proyek_id))
        except Exception as e:
            conn.rollback()
            flash(f'Gagal menyimpan laporan: {str(e)}', 'danger')
            return redirect(url_for('input_laporan', proyek_id=proyek_id))
        finally:
            cursor.close()
            conn.close()

@app.route('/edit_laporan/<int:id>/<int:proyek_id>')
@login_required
def edit_laporan(id, proyek_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    proyek = cursor.fetchone()
    
    cursor.execute("SELECT * FROM laporan_harian WHERE id = %s", (id,))
    laporan = cursor.fetchone()
    
    cursor.execute("SELECT * FROM tenaga_kerja WHERE laporan_id = %s", (id,))
    pekerja = cursor.fetchall()
    
    cursor.execute("SELECT * FROM peralatan WHERE laporan_id = %s", (id,))
    peralatan = cursor.fetchall()
    
    cursor.execute("SELECT * FROM material WHERE laporan_id = %s", (id,))
    material = cursor.fetchall()
    
    cursor.execute("SELECT * FROM pekerjaan WHERE laporan_id = %s", (id,))
    pekerjaan = cursor.fetchall()
    
    cursor.execute("SELECT * FROM kondisi_lapangan WHERE laporan_id = %s", (id,))
    kondisi = cursor.fetchone()
    
    cursor.execute("SELECT * FROM pengesahan WHERE laporan_id = %s", (id,))
    pengesahan = cursor.fetchone()
    
    # Ambil data WBS untuk dropdown jenis pekerjaan
    cursor.execute("SELECT nama_pekerjaan FROM master_wbs WHERE proyek_id = %s ORDER BY kode_wbs ASC", (proyek_id,))
    wbs_list = [r['nama_pekerjaan'] for r in cursor.fetchall()]
    
    # Hitung M1/M2/M3 dari data laporan aktual
    wbs_target_bulanan = _hitung_kumulatif_bulanan(cursor, proyek_id)
    
    conn.close()
    
    if not laporan:
        flash('Laporan tidak ditemukan!', 'danger')
        return redirect(url_for('workspace_history', proyek_id=proyek_id))
    
    g.wbs_list_cache = wbs_list
    return render_template('edit_laporan.html', proyek=proyek, laporan=laporan, pekerja=pekerja, peralatan=peralatan, material=material, pekerjaan=pekerjaan, kondisi=kondisi, pengesahan=pengesahan, wbs_list=wbs_list, wbs_target_bulanan=wbs_target_bulanan)

@app.route('/update_laporan', methods=['POST'])
@login_required
def update_laporan():
    if not validate_csrf(): return redirect(url_for('dashboard'))
    # ── Permission Check ──
    if not can_user_edit(session['user_id']):
        flash('Anda tidak memiliki akses untuk mengedit data!', 'danger')
        return redirect(url_for('dashboard'))
        
    laporan_id = request.form['laporan_id']
    proyek_id = request.form['proyek_id']
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("UPDATE laporan_harian SET tanggal_laporan = %s, cuaca = %s WHERE id = %s",
                       (request.form['tanggal_laporan'], request.form['cuaca'], laporan_id))
        
        cursor.execute("DELETE FROM tenaga_kerja WHERE laporan_id = %s", (laporan_id,))
        cursor.execute("DELETE FROM peralatan WHERE laporan_id = %s", (laporan_id,))
        cursor.execute("DELETE FROM material WHERE laporan_id = %s", (laporan_id,))
        cursor.execute("DELETE FROM pekerjaan WHERE laporan_id = %s", (laporan_id,))
        
        jenis_pekerja = request.form.getlist('jenis_pekerja[]')
        for i in range(len(jenis_pekerja)):
            cursor.execute("INSERT INTO tenaga_kerja (laporan_id, jenis_pekerja, jumlah, hadir, tidak_hadir, keterangan) VALUES (%s, %s, %s, %s, %s, %s)",
                (laporan_id, jenis_pekerja[i], request.form.getlist('jumlah_pekerja[]')[i], request.form.getlist('hadir[]')[i], request.form.getlist('tidak_hadir[]')[i], request.form.getlist('ket_pekerja[]')[i]))
        
        nama_alat = request.form.getlist('nama_alat[]')
        for i in range(len(nama_alat)):
            if nama_alat[i].strip() != '':
                cursor.execute("INSERT INTO peralatan (laporan_id, nama_alat, jumlah_alat, kondisi, keterangan) VALUES (%s, %s, %s, %s, %s)",
                    (laporan_id, nama_alat[i], request.form.getlist('jumlah_alat[]')[i], request.form.getlist('kondisi_alat[]')[i], request.form.getlist('ket_alat[]')[i]))
        
        nama_material = request.form.getlist('nama_material[]')
        for i in range(len(nama_material)):
            if nama_material[i].strip() != '':
                cursor.execute("INSERT INTO material (laporan_id, nama_material, volume_datang, satuan, keterangan) VALUES (%s, %s, %s, %s, %s)",
                    (laporan_id, nama_material[i], request.form.getlist('volume_material[]')[i], request.form.getlist('satuan_material[]')[i], request.form.getlist('ket_material[]')[i]))
        
        jenis_pekerjaan = request.form.getlist('jenis_pekerjaan[]')
        for i in range(len(jenis_pekerjaan)):
            cursor.execute("INSERT INTO pekerjaan (laporan_id, jenis_pekerjaan, lokasi, vol_harian, proses_kumulatif, target_harian, keterangan) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                (laporan_id, jenis_pekerjaan[i], request.form.getlist('lokasi_pekerjaan[]')[i], request.form.getlist('vol_harian[]')[i], request.form.getlist('proses_kumulatif[]')[i], request.form.getlist('target_harian[]')[i], request.form.getlist('ket_pekerjaan[]')[i]))
        
        cursor.execute("UPDATE kondisi_lapangan SET akses = %s, k3 = %s, kondisi_fisik = %s, hambatan = %s, tak_terencana = %s WHERE laporan_id = %s",
            (request.form['akses_lapangan'], request.form['keselamatan'], request.form['kondisi_lapangan'], request.form['faktor_penghambat'], request.form['kegiatan_tak_terencana'], laporan_id))
        
        cursor.execute("UPDATE pengesahan SET nama_pembuat = %s, nama_penyetuju = %s WHERE laporan_id = %s",
            (request.form['nama_pembuat'], request.form['nama_penyetuju'], laporan_id))
        
        conn.commit()
        flash('Laporan Harian Berhasil Diperbarui!', 'success')
        return redirect(url_for('workspace_history', proyek_id=proyek_id))
    except Exception as e:
        conn.rollback()
        flash(f'Gagal memperbarui laporan: {str(e)}', 'danger')
        return redirect(url_for('edit_laporan', id=laporan_id, proyek_id=proyek_id))
    finally:
        cursor.close()
        conn.close()

# ==========================================
# 3. CETAK & UPLOAD DOKUMEN
# ==========================================
@app.route('/cetak/<int:id>')
@login_required
def cetak_laporan(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    sql_info = "SELECT l.*, p.* FROM laporan_harian l JOIN master_proyek p ON l.proyek_id = p.id WHERE l.id = %s"
    cursor.execute(sql_info, (id,))
    info = cursor.fetchone()
    if not info:
        flash('Laporan tidak ditemukan!', 'danger')
        return redirect(url_for('dashboard'))
    cursor.execute("SELECT * FROM tenaga_kerja WHERE laporan_id = %s", (id,))
    pekerja = cursor.fetchall()
    cursor.execute("SELECT * FROM peralatan WHERE laporan_id = %s", (id,))
    peralatan = cursor.fetchall()
    cursor.execute("SELECT * FROM material WHERE laporan_id = %s", (id,))
    material = cursor.fetchall()
    cursor.execute("SELECT * FROM pekerjaan WHERE laporan_id = %s", (id,))
    pekerjaan = cursor.fetchall()
    cursor.execute("SELECT * FROM kondisi_lapangan WHERE laporan_id = %s", (id,))
    kondisi = cursor.fetchone()
    cursor.execute("SELECT * FROM pengesahan WHERE laporan_id = %s", (id,))
    pengesahan = cursor.fetchone()
    conn.close()
    return render_template('laporan.html', info=info, pekerja=pekerja, peralatan=peralatan, material=material, pekerjaan=pekerjaan, kondisi=kondisi, pengesahan=pengesahan)

@app.route('/export_daily_excel/<int:id>')
@login_required
def export_daily_excel(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT l.tanggal_laporan, l.cuaca, l.dokumen_upload,
               p.nama_kegiatan, p.nama_proyek, p.penyedia_jasa, p.konsultan,
               p.pemilik_proyek, p.lokasi, p.no_kontrak, p.tgl_mulai
        FROM laporan_harian l
        JOIN master_proyek p ON l.proyek_id = p.id
        WHERE l.id = %s
    """, (id,))
    info = cursor.fetchone()
    if not info:
        flash('Laporan tidak ditemukan!', 'danger')
        return redirect(url_for('dashboard'))
    cursor.execute("SELECT * FROM tenaga_kerja WHERE laporan_id = %s", (id,))
    pekerja = cursor.fetchall()
    cursor.execute("SELECT * FROM peralatan WHERE laporan_id = %s", (id,))
    peralatan = cursor.fetchall()
    cursor.execute("SELECT * FROM material WHERE laporan_id = %s", (id,))
    material = cursor.fetchall()
    cursor.execute("SELECT * FROM pekerjaan WHERE laporan_id = %s", (id,))
    pekerjaan = cursor.fetchall()
    cursor.execute("SELECT * FROM kondisi_lapangan WHERE laporan_id = %s", (id,))
    kondisi = cursor.fetchone()
    cursor.execute("SELECT * FROM pengesahan WHERE laporan_id = %s", (id,))
    pengesahan = cursor.fetchone()
    conn.close()

    # ── Style Definitions ──
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Harian"

    # Color palette
    DARK_BLUE   = "1B2A4A"
    MED_BLUE    = "2E5090"
    LIGHT_BLUE  = "D6E4F0"
    ACCENT_BLUE = "4472C4"
    WHITE       = "FFFFFF"
    LIGHT_GRAY  = "F2F2F2"
    BLACK       = "000000"
    DARK_GRAY   = "333333"

    # Fonts
    f_title      = Font(name='Calibri', bold=True, size=16, color=WHITE)
    f_subtitle   = Font(name='Calibri', bold=True, size=11, color=WHITE)
    f_label      = Font(name='Calibri', bold=True, size=11, color=DARK_GRAY)
    f_value      = Font(name='Calibri', size=11, color=BLACK)
    f_section    = Font(name='Calibri', bold=True, size=11, color=WHITE)
    f_header     = Font(name='Calibri', bold=True, size=10, color=WHITE)
    f_data       = Font(name='Calibri', size=10, color=BLACK)
    f_data_bold  = Font(name='Calibri', bold=True, size=10, color=BLACK)
    f_center     = Font(name='Calibri', size=10, color=BLACK)
    f_note       = Font(name='Calibri', italic=True, size=10, color="666666")

    # Fills
    fill_title    = PatternFill('solid', fgColor=DARK_BLUE)
    fill_section  = PatternFill('solid', fgColor=MED_BLUE)
    fill_header   = PatternFill('solid', fgColor=ACCENT_BLUE)
    fill_label_bg = PatternFill('solid', fgColor=LIGHT_BLUE)
    fill_alt_row  = PatternFill('solid', fgColor=LIGHT_GRAY)
    fill_white    = PatternFill('solid', fgColor=WHITE)

    # Borders
    thin  = Side(style='thin', color='B0B0B0')
    med   = Side(style='medium', color=MED_BLUE)
    b_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    b_sec = Border(left=med, right=med, top=med, bottom=med)

    # Alignments
    al_left   = Alignment(horizontal='left', vertical='center', wrap_text=True)
    al_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al_right  = Alignment(horizontal='right', vertical='center', wrap_text=True)

    # Column widths (8 columns max)
    col_widths = {'A': 6, 'B': 30, 'C': 18, 'D': 18, 'E': 16, 'F': 14, 'G': 14, 'H': 25}
    for letter, w in col_widths.items():
        ws.column_dimensions[letter].width = w

    # ── Helper functions ──
    def set_cell(r, c, val, font=f_data, fill=fill_white, align=al_left, border=b_all):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = font
        cell.fill = fill
        cell.alignment = align
        cell.border = border
        return cell

    def section_banner(r, title, col_span=8):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=col_span)
        c = ws.cell(row=r, column=1, value=title)
        c.font = f_section
        c.fill = fill_section
        c.alignment = Alignment(horizontal='left', vertical='center')
        for ci in range(1, col_span + 1):
            ws.cell(row=r, column=ci).border = b_sec
            ws.cell(row=r, column=ci).fill = fill_section
        ws.row_dimensions[r].height = 26
        return r + 1

    def table_header(r, headers, col_span=None):
        if col_span is None:
            col_span = len(headers)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font = f_header
            c.fill = fill_header
            c.alignment = al_center
            c.border = b_all
        ws.row_dimensions[r].height = 22
        return r + 1

    row = 1

    # ══════════════════════════════════════════
    # BANNER JUDUL
    # ══════════════════════════════════════════
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value="LAPORAN HARIAN PELAKSANAAN PEKERJAAN KONSTRUKSI")
    c.font = f_title
    c.fill = fill_title
    c.alignment = Alignment(horizontal='center', vertical='center')
    for ci in range(1, 9):
        ws.cell(row=row, column=ci).fill = fill_title
        ws.cell(row=row, column=ci).border = b_sec
    ws.row_dimensions[row].height = 35
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value=str(info.get('nama_proyek', '')))
    c.font = f_subtitle
    c.fill = fill_title
    c.alignment = Alignment(horizontal='center', vertical='center')
    for ci in range(1, 9):
        ws.cell(row=row, column=ci).fill = fill_title
        ws.cell(row=row, column=ci).border = b_sec
    ws.row_dimensions[row].height = 24
    row += 1

    # No. Kontrak & Lokasi baris
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
    set_cell(row, 1, f"No. Kontrak: {info.get('no_kontrak', '-')}", f_subtitle, fill_title, al_left)
    set_cell(row, 5, f"Lokasi: {info.get('lokasi', '-')}", f_subtitle, fill_title, al_left)
    for ci in range(1, 9):
        ws.cell(row=row, column=ci).fill = fill_title
        ws.cell(row=row, column=ci).border = b_sec
    ws.row_dimensions[row].height = 20
    row += 1

    # ══════════════════════════════════════════
    # INFO PROYEK (2 kolom side-by-side)
    # ══════════════════════════════════════════
    row += 1  # spacer

    # Left column: info proyek
    left_labels = [
        ("Hari / Tanggal", str(info.get('tanggal_laporan', '-'))),
        ("Kondisi Cuaca", str(info.get('cuaca', '-'))),
        ("Pemilik Proyek", str(info.get('pemilik_proyek', '-'))),
    ]
    right_labels = [
        ("Penyedia Jasa", str(info.get('penyedia_jasa', '-'))),
        ("Konsultan Pengawas", str(info.get('konsultan', '-'))),
        ("Tanggal Mulai", str(info.get('tgl_mulai', '-'))),
    ]
    for i in range(3):
        # Left: cols A-B
        set_cell(row, 1, left_labels[i][0], f_label, fill_label_bg, al_left, b_all)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        set_cell(row, 2, left_labels[i][1], f_value, fill_white, al_left, b_all)
        for ci in [3, 4]:
            ws.cell(row=row, column=ci).border = b_all
        # Right: cols E-H
        set_cell(row, 5, right_labels[i][0], f_label, fill_label_bg, al_left, b_all)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
        set_cell(row, 6, right_labels[i][1], f_value, fill_white, al_left, b_all)
        for ci in [7, 8]:
            ws.cell(row=row, column=ci).border = b_all
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1  # spacer

    # ══════════════════════════════════════════
    # 1. TENAGA KERJA
    # ══════════════════════════════════════════
    row = section_banner(row, "I. TENAGA KERJA / PERSONIL")
    tk_headers = ["No", "Jenis Pekerja / Jabatan", "Jumlah", "Hadir", "Tidak Hadir", "Keterangan", "", ""]
    row = table_header(row, tk_headers)
    for i, p in enumerate(pekerja, 1):
        alt = fill_alt_row if i % 2 == 0 else fill_white
        set_cell(row, 1, i, f_center, alt, al_center)
        set_cell(row, 2, p.get('jenis_pekerja', ''), f_data, alt, al_left)
        set_cell(row, 3, p.get('jumlah', 0), f_data, alt, al_center)
        set_cell(row, 4, p.get('hadir', 0), f_data, alt, al_center)
        set_cell(row, 5, p.get('tidak_hadir', 0), f_data, alt, al_center)
        set_cell(row, 6, p.get('keterangan', ''), f_data, alt, al_left)
        set_cell(row, 7, '', f_data, alt, al_center)
        set_cell(row, 8, '', f_data, alt, al_center)
        ws.row_dimensions[row].height = 20
        row += 1
    if not pekerja:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        set_cell(row, 1, "Tidak ada data tenaga kerja.", f_note, fill_white, al_center)
        for ci in range(2, 9):
            ws.cell(row=row, column=ci).border = b_all
        row += 1
    row += 1  # spacer

    # ══════════════════════════════════════════
    # 2. PERALATAN
    # ══════════════════════════════════════════
    row = section_banner(row, "II. PERALATAN / ALAT BERAT")
    al_headers = ["No", "Nama Alat / Jenis Peralatan", "Jumlah Unit", "Kondisi", "Keterangan", "", "", ""]
    row = table_header(row, al_headers)
    for i, a in enumerate(peralatan, 1):
        alt = fill_alt_row if i % 2 == 0 else fill_white
        set_cell(row, 1, i, f_center, alt, al_center)
        set_cell(row, 2, a.get('nama_alat', ''), f_data, alt, al_left)
        set_cell(row, 3, a.get('jumlah_alat', 0), f_data, alt, al_center)
        set_cell(row, 4, a.get('kondisi', ''), f_data, alt, al_center)
        set_cell(row, 5, a.get('keterangan', ''), f_data, alt, al_left)
        set_cell(row, 6, '', f_data, alt, al_center)
        set_cell(row, 7, '', f_data, alt, al_center)
        set_cell(row, 8, '', f_data, alt, al_center)
        ws.row_dimensions[row].height = 20
        row += 1
    if not peralatan:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        set_cell(row, 1, "Tidak ada data peralatan.", f_note, fill_white, al_center)
        for ci in range(2, 9):
            ws.cell(row=row, column=ci).border = b_all
        row += 1
    row += 1  # spacer

    # ══════════════════════════════════════════
    # 3. MATERIAL
    # ══════════════════════════════════════════
    row = section_banner(row, "III. MATERIAL / BAHAN DATANG")
    mt_headers = ["No", "Nama Material / Bahan", "Volume Datang", "Satuan", "Keterangan", "", "", ""]
    row = table_header(row, mt_headers)
    for i, m in enumerate(material, 1):
        alt = fill_alt_row if i % 2 == 0 else fill_white
        set_cell(row, 1, i, f_center, alt, al_center)
        set_cell(row, 2, m.get('nama_material', ''), f_data, alt, al_left)
        set_cell(row, 3, m.get('volume_datang', 0), f_data, alt, al_center)
        set_cell(row, 4, m.get('satuan', ''), f_data, alt, al_center)
        set_cell(row, 5, m.get('keterangan', ''), f_data, alt, al_left)
        set_cell(row, 6, '', f_data, alt, al_center)
        set_cell(row, 7, '', f_data, alt, al_center)
        set_cell(row, 8, '', f_data, alt, al_center)
        ws.row_dimensions[row].height = 20
        row += 1
    if not material:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        set_cell(row, 1, "Tidak ada data material.", f_note, fill_white, al_center)
        for ci in range(2, 9):
            ws.cell(row=row, column=ci).border = b_all
        row += 1
    row += 1  # spacer

    # ══════════════════════════════════════════
    # 4. PEKERJAAN / PROGRES
    # ══════════════════════════════════════════
    row = section_banner(row, "IV. PEKERJAAN / PROGRES LAPANGAN")
    pk_headers = ["No", "Jenis Pekerjaan", "Lokasi", "Vol Harian", "Kumulatif (%)", "Target (%)", "Keterangan", ""]
    row = table_header(row, pk_headers)
    for i, pk in enumerate(pekerjaan, 1):
        alt = fill_alt_row if i % 2 == 0 else fill_white
        set_cell(row, 1, i, f_center, alt, al_center)
        set_cell(row, 2, pk.get('jenis_pekerjaan', ''), f_data, alt, al_left)
        set_cell(row, 3, pk.get('lokasi', ''), f_data, alt, al_left)
        set_cell(row, 4, pk.get('vol_harian', ''), f_data, alt, al_center)
        set_cell(row, 5, pk.get('proses_kumulatif', ''), f_data_bold, alt, al_center)
        set_cell(row, 6, pk.get('target_harian', ''), f_data, alt, al_center)
        set_cell(row, 7, pk.get('keterangan', ''), f_data, alt, al_left)
        set_cell(row, 8, '', f_data, alt, al_center)
        ws.row_dimensions[row].height = 20
        row += 1
    if not pekerjaan:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        set_cell(row, 1, "Tidak ada data pekerjaan.", f_note, fill_white, al_center)
        for ci in range(2, 9):
            ws.cell(row=row, column=ci).border = b_all
        row += 1
    row += 1  # spacer

    # ══════════════════════════════════════════
    # 5. KONDISI LAPANGAN
    # ══════════════════════════════════════════
    row = section_banner(row, "V. KONDISI LAPANGAN & CATATAN")
    kondisi_fields = [
        ("Akses Lapangan", 'akses'),
        ("Keselamatan Kerja (K3)", 'k3'),
        ("Kondisi Fisik Lapangan", 'kondisi_fisik'),
        ("Faktor Penghambat / Kendala", 'hambatan'),
        ("Kegiatan Tak Terencana", 'tak_terencana'),
    ]
    for i, (label, key) in enumerate(kondisi_fields):
        alt = fill_alt_row if i % 2 == 0 else fill_white
        set_cell(row, 1, label, f_label, fill_label_bg, al_left, b_all)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        val = str(kondisi.get(key, '-')) if kondisi else '-'
        set_cell(row, 2, val, f_data, alt, al_left, b_all)
        for ci in range(3, 9):
            ws.cell(row=row, column=ci).border = b_all
            ws.cell(row=row, column=ci).fill = alt
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1  # spacer

    # ══════════════════════════════════════════
    # 6. PENGESAHAN
    # ══════════════════════════════════════════
    row = section_banner(row, "VI. PENGESAHAN LAPORAN")
    row += 1  # spacer

    # Dibuat oleh
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
    set_cell(row, 1, "Dibuat Oleh:", f_label, fill_label_bg, al_center, b_all)
    set_cell(row, 5, "Disetujui Oleh:", f_label, fill_label_bg, al_center, b_all)
    for ci in [2, 3, 4, 6, 7, 8]:
        ws.cell(row=row, column=ci).border = b_all
    ws.row_dimensions[row].height = 20
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
    set_cell(row, 1, "Pelaksana / Site Manager", f_note, fill_white, al_center, b_all)
    set_cell(row, 5, "Konsultan Pengawas / Owner", f_note, fill_white, al_center, b_all)
    for ci in [2, 3, 4, 6, 7, 8]:
        ws.cell(row=row, column=ci).border = b_all
    ws.row_dimensions[row].height = 18
    row += 2  # space for signature

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
    nama_pembuat = str(pengesahan.get('nama_pembuat', '')) if pengesahan else ''
    nama_penyetuju = str(pengesahan.get('nama_penyetuju', '')) if pengesahan else ''
    set_cell(row, 1, f"({nama_pembuat})", f_data_bold, fill_white, al_center, Border(bottom=Side(style='thin')))
    set_cell(row, 5, f"({nama_penyetuju})", f_data_bold, fill_white, al_center, Border(bottom=Side(style='thin')))
    for ci in [2, 3, 4, 6, 7, 8]:
        ws.cell(row=row, column=ci).border = Border(bottom=Side(style='thin'))
    ws.row_dimensions[row].height = 20
    row += 2

    # ══════════════════════════════════════════
    # FOOTER
    # ══════════════════════════════════════════
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value="Dicetak oleh Sistem MANPRO — Software Manajemen Proyek Terpadu")
    c.font = Font(name='Calibri', italic=True, size=9, color='999999')
    c.alignment = Alignment(horizontal='right', vertical='center')

    # Print settings
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f'A1:H{row}'

    filename = f"Laporan_Harian_{info.get('nama_proyek', 'Proyek')}_{info.get('tanggal_laporan', '')}.xlsx"
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/upload/<int:id>', methods=['POST'])
@login_required
def upload_dokumen(id):
    if not validate_csrf(): return redirect(url_for('dashboard'))
    file = request.files.get('file')
    proyek_id = request.form['proyek_id']
    if file and file.filename != '':
        filename = secure_filename(file.filename)
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE laporan_harian SET dokumen_upload = %s WHERE id = %s", (filename, id))
        conn.commit()
        conn.close()
        flash('Dokumen berhasil diupload!', 'success')
    return redirect(url_for('workspace_history', proyek_id=proyek_id))

@app.route('/delete/<int:id>/<int:proyek_id>')
@login_required
def hapus_laporan(id, proyek_id):
    # ── Permission Check ──
    if not can_user_delete(session['user_id']):
        flash('Anda tidak memiliki akses untuk menghapus data!', 'danger')
        return redirect(url_for('workspace_history', proyek_id=proyek_id))
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM laporan_harian WHERE id = %s", (id,))
    conn.commit()
    conn.close()
    flash('Laporan berhasil dihapus!', 'danger')
    return redirect(url_for('workspace_history', proyek_id=proyek_id))

# ==========================================
# 4. WEEKLY & MONTHLY REPORT + CETAK PDF
# ==========================================
@app.route('/weekly/<int:proyek_id>')
@login_required
def weekly_report(proyek_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    proyek = cursor.fetchone()
    sql_weekly = """
        SELECT minggu, MIN(tanggal_laporan) as awal_minggu, MAX(tanggal_laporan) as akhir_minggu,
               COUNT(DISTINCT id) as total_hari_kerja, SUM(hadir) as total_pekerja_hadir
        FROM (
            SELECT lh.id, lh.tanggal_laporan, YEARWEEK(lh.tanggal_laporan, 1) as minggu, tk.hadir
            FROM laporan_harian lh LEFT JOIN tenaga_kerja tk ON lh.id = tk.laporan_id
            WHERE lh.proyek_id = %s
        ) sub
        GROUP BY minggu ORDER BY minggu DESC
    """
    cursor.execute(sql_weekly, (proyek_id,))
    mingguan = cursor.fetchall()
    conn.close()
    return render_template('weekly_report.html', proyek=proyek, mingguan=mingguan)

@app.route('/cetak_weekly/<int:proyek_id>/<string:minggu>')
@login_required
def cetak_weekly(proyek_id, minggu):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    info = cursor.fetchone()
    sql = """SELECT lh.*, p.jenis_pekerjaan, p.vol_harian, p.proses_kumulatif
             FROM laporan_harian lh LEFT JOIN pekerjaan p ON lh.id = p.laporan_id
             WHERE lh.proyek_id = %s AND YEARWEEK(lh.tanggal_laporan, 1) = %s ORDER BY lh.tanggal_laporan ASC"""
    cursor.execute(sql, (proyek_id, minggu))
    detail = cursor.fetchall()
    conn.close()
    return render_template('cetak_weekly.html', info=info, detail=detail, minggu=minggu)

@app.route('/export_weekly_excel/<int:proyek_id>/<string:minggu>')
@login_required
def export_weekly_excel(proyek_id, minggu):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    info = cursor.fetchone()
    sql = """SELECT lh.*, p.jenis_pekerjaan, p.vol_harian, p.proses_kumulatif
             FROM laporan_harian lh LEFT JOIN pekerjaan p ON lh.id = p.laporan_id
             WHERE lh.proyek_id = %s AND YEARWEEK(lh.tanggal_laporan, 1) = %s ORDER BY lh.tanggal_laporan ASC"""
    cursor.execute(sql, (proyek_id, minggu))
    detail = cursor.fetchall()

    # Hitung statistik
    total_hari = len(set(str(d.get('tanggal_laporan','')) for d in detail))
    total_pekerja_hadir = sum(d.get('hadir', 0) or 0 for d in detail if d.get('jenis_pekerja'))
    # Ambil progres terakhir
    progres_akhir = '-'
    for d in reversed(detail):
        if d.get('proses_kumulatif'):
            progres_akhir = d['proses_kumulatif']
            break
    conn.close()

    # ── Style Definitions ──
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Report"

    DARK_BLUE   = "1B2A4A"
    MED_BLUE    = "2E5090"
    ACCENT_BLUE = "4472C4"
    WHITE       = "FFFFFF"
    LIGHT_GRAY  = "F2F2F2"
    BLACK       = "000000"
    DARK_GRAY   = "333333"
    LIGHT_BLUE  = "D6E4F0"

    f_title    = Font(name='Calibri', bold=True, size=16, color=WHITE)
    f_subtitle = Font(name='Calibri', bold=True, size=11, color=WHITE)
    f_label    = Font(name='Calibri', bold=True, size=11, color=DARK_GRAY)
    f_value    = Font(name='Calibri', size=11, color=BLACK)
    f_section  = Font(name='Calibri', bold=True, size=11, color=WHITE)
    f_header   = Font(name='Calibri', bold=True, size=10, color=WHITE)
    f_data     = Font(name='Calibri', size=10, color=BLACK)
    f_data_c   = Font(name='Calibri', size=10, color=BLACK)
    f_bold     = Font(name='Calibri', bold=True, size=10, color=BLACK)
    f_note     = Font(name='Calibri', italic=True, size=10, color="666666")
    f_stat_lbl = Font(name='Calibri', bold=True, size=10, color=WHITE)
    f_stat_val = Font(name='Calibri', bold=True, size=14, color=WHITE)

    fill_title   = PatternFill('solid', fgColor=DARK_BLUE)
    fill_section = PatternFill('solid', fgColor=MED_BLUE)
    fill_header  = PatternFill('solid', fgColor=ACCENT_BLUE)
    fill_label   = PatternFill('solid', fgColor=LIGHT_BLUE)
    fill_alt     = PatternFill('solid', fgColor=LIGHT_GRAY)
    fill_white   = PatternFill('solid', fgColor=WHITE)
    fill_green   = PatternFill('solid', fgColor="27AE60")
    fill_orange  = PatternFill('solid', fgColor="E67E22")

    thin  = Side(style='thin', color='B0B0B0')
    med   = Side(style='medium', color=MED_BLUE)
    b_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    b_sec = Border(left=med, right=med, top=med, bottom=med)

    al_left   = Alignment(horizontal='left', vertical='center', wrap_text=True)
    al_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al_right  = Alignment(horizontal='right', vertical='center', wrap_text=True)

    col_widths = {'A': 6, 'B': 16, 'C': 16, 'D': 32, 'E': 18, 'F': 18}
    for letter, w in col_widths.items():
        ws.column_dimensions[letter].width = w

    def set_cell(r, c, val, font=f_data, fill=fill_white, align=al_left, border=b_all):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = font
        cell.fill = fill
        cell.alignment = align
        cell.border = border
        return cell

    row = 1  
    # ══════════════════════════════════════════
    # BANNER JUDUL
    # ══════════════════════════════════════════
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value="LAPORAN REKAPITULASI MINGGUAN")
    c.font = f_title; c.fill = fill_title
    c.alignment = Alignment(horizontal='center', vertical='center')
    for ci in range(1, 7):
        ws.cell(row=row, column=ci).fill = fill_title
        ws.cell(row=row, column=ci).border = b_sec
    ws.row_dimensions[row].height = 35
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value=str(info.get('nama_proyek', '')))
    c.font = f_subtitle; c.fill = fill_title
    c.alignment = Alignment(horizontal='center', vertical='center')
    for ci in range(1, 7):
        ws.cell(row=row, column=ci).fill = fill_title
        ws.cell(row=row, column=ci).border = b_sec
    ws.row_dimensions[row].height = 24
    row += 1

    # No. Kontrak & Lokasi
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    set_cell(row, 1, f"No. Kontrak: {info.get('no_kontrak', '-')}", f_subtitle, fill_title, al_left)
    set_cell(row, 4, f"Lokasi: {info.get('lokasi', '-')}", f_subtitle, fill_title, al_left)
    for ci in range(1, 7):
        ws.cell(row=row, column=ci).fill = fill_title
        ws.cell(row=row, column=ci).border = b_sec
    ws.row_dimensions[row].height = 20
    row += 1

    # ══════════════════════════════════════════
    # INFO PROYEK
    # ══════════════════════════════════════════
    row += 1
    info_left = [
        ("Minggu Ke-", str(minggu)),
        ("Penyedia Jasa", str(info.get('penyedia_jasa', '-'))),
    ]
    info_right = [
        ("Konsultan Pengawas", str(info.get('konsultan', '-'))),
        ("Pemilik Proyek", str(info.get('pemilik_proyek', '-'))),
    ]
    for i in range(2):
        set_cell(row, 1, info_left[i][0], f_label, fill_label, al_left)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        set_cell(row, 2, info_left[i][1], f_value, fill_white, al_left)
        ws.cell(row=row, column=3).border = b_all
        set_cell(row, 4, info_right[i][0], f_label, fill_label, al_left)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        set_cell(row, 5, info_right[i][1], f_value, fill_white, al_left)
        ws.cell(row=row, column=6).border = b_all
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1

    # ══════════════════════════════════════════
    # STATISTIK RINGKASAN
    # ══════════════════════════════════════════
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value="RINGKASAN MINGGU INI")
    c.font = f_section; c.fill = fill_section
    c.alignment = Alignment(horizontal='left', vertical='center')
    for ci in range(1, 7):
        ws.cell(row=row, column=ci).fill = fill_section
        ws.cell(row=row, column=ci).border = b_sec
    ws.row_dimensions[row].height = 26
    row += 1

    # 3 stat cards
    stats = [
        ("Total Hari Kerja", str(total_hari), fill_green),
        ("Total Pekerja Hadir", str(total_pekerja_hadir), fill_orange),
        ("Progres Akhir", str(progres_akhir), PatternFill('solid', fgColor=MED_BLUE)),
    ]
    for ci, (label, val, fill) in enumerate(stats):
        col_start = ci * 2 + 1
        col_end = col_start + 1
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        c = ws.cell(row=row, column=col_start, value=label)
        c.font = f_stat_lbl; c.fill = fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=col_end).fill = fill
        for x in [col_start, col_end]:
            ws.cell(row=row, column=x).border = b_sec
    ws.row_dimensions[row].height = 22
    row += 1
    for ci, (label, val, fill) in enumerate(stats):
        col_start = ci * 2 + 1
        col_end = col_start + 1
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        c = ws.cell(row=row, column=col_start, value=val)
        c.font = f_stat_val; c.fill = fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=col_end).fill = fill
        for x in [col_start, col_end]:
            ws.cell(row=row, column=x).border = b_sec
    ws.row_dimensions[row].height = 30
    row += 2

    # ══════════════════════════════════════════
    # TABEL DETAIL
    # ══════════════════════════════════════════
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value="DETAIL LAPORAN HARIAN")
    c.font = f_section; c.fill = fill_section
    c.alignment = Alignment(horizontal='left', vertical='center')
    for ci in range(1, 7):
        ws.cell(row=row, column=ci).fill = fill_section
        ws.cell(row=row, column=ci).border = b_sec
    ws.row_dimensions[row].height = 26
    row += 1

    headers = ["No", "Tanggal", "Cuaca", "Jenis Pekerjaan", "Volume Harian", "Progres Kumulatif"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = f_header; c.fill = fill_header
        c.alignment = al_center; c.border = b_all
    ws.row_dimensions[row].height = 22
    row += 1

    for i, d in enumerate(detail, 1):
        alt = fill_alt if i % 2 == 0 else fill_white
        set_cell(row, 1, i, f_data_c, alt, al_center)
        set_cell(row, 2, str(d.get('tanggal_laporan', '')), f_data, alt, al_center)
        set_cell(row, 3, d.get('cuaca', ''), f_data, alt, al_center)
        set_cell(row, 4, d.get('jenis_pekerjaan', '') or '-', f_data, alt, al_left)
        set_cell(row, 5, d.get('vol_harian', '') or '-', f_data, alt, al_center)
        set_cell(row, 6, d.get('proses_kumulatif', '') or '-', f_bold, alt, al_center)
        ws.row_dimensions[row].height = 20
        row += 1

    if not detail:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        set_cell(row, 1, "Tidak ada data laporan mingguan.", f_note, fill_white, al_center)
        for ci in range(2, 7):
            ws.cell(row=row, column=ci).border = b_all
        row += 1

    # ══════════════════════════════════════════
    # FOOTER
    # ══════════════════════════════════════════
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value="Dicetak oleh Sistem MANPRO — Software Manajemen Proyek Terpadu")
    c.font = Font(name='Calibri', italic=True, size=9, color='999999')
    c.alignment = Alignment(horizontal='right', vertical='center')

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f'A1:F{row}'

    filename = f"Weekly_Report_{info.get('nama_proyek','')}_Minggu_{minggu}.xlsx"
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/monthly/<int:proyek_id>')
@login_required
def monthly_report(proyek_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    proyek = cursor.fetchone()
    sql_monthly = """
        SELECT DATE_FORMAT(lh.tanggal_laporan, '%Y-%m') as bulan, DATE_FORMAT(lh.tanggal_laporan, '%M %Y') as nama_bulan,
               COUNT(DISTINCT lh.id) as total_hari_kerja, SUM(tk.hadir) as total_pekerja_hadir
        FROM laporan_harian lh LEFT JOIN tenaga_kerja tk ON lh.id = tk.laporan_id
        WHERE lh.proyek_id = %s GROUP BY bulan, nama_bulan ORDER BY bulan DESC
    """
    cursor.execute(sql_monthly, (proyek_id,))
    bulanan = cursor.fetchall()
    conn.close()
    return render_template('monthly_report.html',
                            proyek=proyek,
                            bulanan=bulanan)

@app.route('/cetak_monthly/<int:proyek_id>/<string:bulan>')
@login_required
def cetak_monthly(proyek_id, bulan):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    info = cursor.fetchone()
    sql = """SELECT lh.*, p.jenis_pekerjaan, p.vol_harian, p.proses_kumulatif
             FROM laporan_harian lh LEFT JOIN pekerjaan p ON lh.id = p.laporan_id
             WHERE lh.proyek_id = %s AND DATE_FORMAT(lh.tanggal_laporan, '%Y-%m') = %s ORDER BY lh.tanggal_laporan ASC"""
    cursor.execute(sql, (proyek_id, bulan))
    detail = cursor.fetchall()
    conn.close()
    return render_template('cetak_monthly.html', info=info, detail=detail, bulan=bulan)

@app.route('/export_monthly_excel/<int:proyek_id>/<string:bulan>')
@login_required
def export_monthly_excel(proyek_id, bulan):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    info = cursor.fetchone()
    sql = """SELECT lh.*, p.jenis_pekerjaan, p.vol_harian, p.proses_kumulatif
             FROM laporan_harian lh LEFT JOIN pekerjaan p ON lh.id = p.laporan_id
             WHERE lh.proyek_id = %s AND DATE_FORMAT(lh.tanggal_laporan, %s) = %s ORDER BY lh.tanggal_laporan ASC"""
    cursor.execute(sql, (proyek_id, '%Y-%m', bulan))
    detail = cursor.fetchall()

    # Hitung statistik
    total_hari = len(set(str(d.get('tanggal_laporan','')) for d in detail))
    total_pekerja_hadir = sum(d.get('hadir', 0) or 0 for d in detail if d.get('jenis_pekerja'))
    progres_awal = '-'
    progres_akhir = '-'
    for d in detail:
        if d.get('proses_kumulatif') and progres_awal == '-':
            progres_awal = d['proses_kumulatif']
    for d in reversed(detail):
        if d.get('proses_kumulatif'):
            progres_akhir = d['proses_kumulatif']
            break
    # Hitung nama bulan
    try:
        from datetime import datetime
        dt = datetime.strptime(bulan, '%Y-%m')
        nama_bulan = dt.strftime('%B %Y')
    except:
        nama_bulan = bulan
    conn.close()

    # ── Style Definitions ──
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Report"

    DARK_BLUE   = "1B2A4A"
    MED_BLUE    = "2E5090"
    ACCENT_BLUE = "4472C4"
    WHITE       = "FFFFFF"
    LIGHT_GRAY  = "F2F2F2"
    BLACK       = "000000"
    DARK_GRAY   = "333333"
    LIGHT_BLUE  = "D6E4F0"

    f_title    = Font(name='Calibri', bold=True, size=16, color=WHITE)
    f_subtitle = Font(name='Calibri', bold=True, size=11, color=WHITE)
    f_label    = Font(name='Calibri', bold=True, size=11, color=DARK_GRAY)
    f_value    = Font(name='Calibri', size=11, color=BLACK)
    f_section  = Font(name='Calibri', bold=True, size=11, color=WHITE)
    f_header   = Font(name='Calibri', bold=True, size=10, color=WHITE)
    f_data     = Font(name='Calibri', size=10, color=BLACK)
    f_data_c   = Font(name='Calibri', size=10, color=BLACK)
    f_bold     = Font(name='Calibri', bold=True, size=10, color=BLACK)
    f_note     = Font(name='Calibri', italic=True, size=10, color="666666")
    f_stat_lbl = Font(name='Calibri', bold=True, size=10, color=WHITE)
    f_stat_val = Font(name='Calibri', bold=True, size=14, color=WHITE)

    fill_title   = PatternFill('solid', fgColor=DARK_BLUE)
    fill_section = PatternFill('solid', fgColor=MED_BLUE)
    fill_header  = PatternFill('solid', fgColor=ACCENT_BLUE)
    fill_label   = PatternFill('solid', fgColor=LIGHT_BLUE)
    fill_alt     = PatternFill('solid', fgColor=LIGHT_GRAY)
    fill_white   = PatternFill('solid', fgColor=WHITE)
    fill_green   = PatternFill('solid', fgColor="27AE60")
    fill_orange  = PatternFill('solid', fgColor="E67E22")

    thin  = Side(style='thin', color='B0B0B0')
    med   = Side(style='medium', color=MED_BLUE)
    b_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    b_sec = Border(left=med, right=med, top=med, bottom=med)

    al_left   = Alignment(horizontal='left', vertical='center', wrap_text=True)
    al_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al_right  = Alignment(horizontal='right', vertical='center', wrap_text=True)

    col_widths = {'A': 6, 'B': 16, 'C': 16, 'D': 32, 'E': 18, 'F': 18}
    for letter, w in col_widths.items():
        ws.column_dimensions[letter].width = w

    def set_cell(r, c, val, font=f_data, fill=fill_white, align=al_left, border=b_all):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = font
        cell.fill = fill
        cell.alignment = align
        cell.border = border
        return cell

    row = 1

    # ══════════════════════════════════════════
    # BANNER JUDUL
    # ══════════════════════════════════════════
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value="LAPORAN REKAPITULASI BULANAN")
    c.font = f_title; c.fill = fill_title
    c.alignment = Alignment(horizontal='center', vertical='center')
    for ci in range(1, 7):
        ws.cell(row=row, column=ci).fill = fill_title
        ws.cell(row=row, column=ci).border = b_sec
    ws.row_dimensions[row].height = 35
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value=str(info.get('nama_proyek', '')))
    c.font = f_subtitle; c.fill = fill_title
    c.alignment = Alignment(horizontal='center', vertical='center')
    for ci in range(1, 7):
        ws.cell(row=row, column=ci).fill = fill_title
        ws.cell(row=row, column=ci).border = b_sec
    ws.row_dimensions[row].height = 24
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    set_cell(row, 1, f"No. Kontrak: {info.get('no_kontrak', '-')}", f_subtitle, fill_title, al_left)
    set_cell(row, 4, f"Lokasi: {info.get('lokasi', '-')}", f_subtitle, fill_title, al_left)
    for ci in range(1, 7):
        ws.cell(row=row, column=ci).fill = fill_title
        ws.cell(row=row, column=ci).border = b_sec
    ws.row_dimensions[row].height = 20
    row += 1

    # ══════════════════════════════════════════
    # INFO PROYEK
    # ══════════════════════════════════════════
    row += 1
    info_left = [
        ("Periode Bulan", nama_bulan),
        ("Penyedia Jasa", str(info.get('penyedia_jasa', '-'))),
    ]
    info_right = [
        ("Konsultan Pengawas", str(info.get('konsultan', '-'))),
        ("Pemilik Proyek", str(info.get('pemilik_proyek', '-'))),
    ]
    for i in range(2):
        set_cell(row, 1, info_left[i][0], f_label, fill_label, al_left)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        set_cell(row, 2, info_left[i][1], f_value, fill_white, al_left)
        ws.cell(row=row, column=3).border = b_all
        set_cell(row, 4, info_right[i][0], f_label, fill_label, al_left)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        set_cell(row, 5, info_right[i][1], f_value, fill_white, al_left)
        ws.cell(row=row, column=6).border = b_all
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1

    # ══════════════════════════════════════════
    # STATISTIK RINGKASAN
    # ══════════════════════════════════════════
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value="RINGKASAN BULAN INI")
    c.font = f_section; c.fill = fill_section
    c.alignment = Alignment(horizontal='left', vertical='center')
    for ci in range(1, 7):
        ws.cell(row=row, column=ci).fill = fill_section
        ws.cell(row=row, column=ci).border = b_sec
    ws.row_dimensions[row].height = 26
    row += 1

    stats = [
        ("Total Hari Kerja", str(total_hari), fill_green),
        ("Total Pekerja Hadir", str(total_pekerja_hadir), fill_orange),
        ("Progres Akhir", str(progres_akhir), PatternFill('solid', fgColor=MED_BLUE)),
    ]
    for ci, (label, val, fill) in enumerate(stats):
        col_start = ci * 2 + 1
        col_end = col_start + 1
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        c = ws.cell(row=row, column=col_start, value=label)
        c.font = f_stat_lbl; c.fill = fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=col_end).fill = fill
        for x in [col_start, col_end]:
            ws.cell(row=row, column=x).border = b_sec
    ws.row_dimensions[row].height = 22
    row += 1
    for ci, (label, val, fill) in enumerate(stats):
        col_start = ci * 2 + 1
        col_end = col_start + 1
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        c = ws.cell(row=row, column=col_start, value=val)
        c.font = f_stat_val; c.fill = fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=col_end).fill = fill
        for x in [col_start, col_end]:
            ws.cell(row=row, column=x).border = b_sec
    ws.row_dimensions[row].height = 30
    row += 2

    # ══════════════════════════════════════════
    # TABEL DETAIL
    # ══════════════════════════════════════════
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value="DETAIL LAPORAN HARIAN")
    c.font = f_section; c.fill = fill_section
    c.alignment = Alignment(horizontal='left', vertical='center')
    for ci in range(1, 7):
        ws.cell(row=row, column=ci).fill = fill_section
        ws.cell(row=row, column=ci).border = b_sec
    ws.row_dimensions[row].height = 26
    row += 1

    headers = ["No", "Tanggal", "Cuaca", "Jenis Pekerjaan", "Volume Harian", "Progres Kumulatif"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = f_header; c.fill = fill_header
        c.alignment = al_center; c.border = b_all
    ws.row_dimensions[row].height = 22
    row += 1

    for i, d in enumerate(detail, 1):
        alt = fill_alt if i % 2 == 0 else fill_white
        set_cell(row, 1, i, f_data_c, alt, al_center)
        set_cell(row, 2, str(d.get('tanggal_laporan', '')), f_data, alt, al_center)
        set_cell(row, 3, d.get('cuaca', ''), f_data, alt, al_center)
        set_cell(row, 4, d.get('jenis_pekerjaan', '') or '-', f_data, alt, al_left)
        set_cell(row, 5, d.get('vol_harian', '') or '-', f_data, alt, al_center)
        set_cell(row, 6, d.get('proses_kumulatif', '') or '-', f_bold, alt, al_center)
        ws.row_dimensions[row].height = 20
        row += 1

    if not detail:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        set_cell(row, 1, "Tidak ada data laporan bulanan.", f_note, fill_white, al_center)
        for ci in range(2, 7):
            ws.cell(row=row, column=ci).border = b_all
        row += 1

    # ══════════════════════════════════════════
    # FOOTER
    # ══════════════════════════════════════════
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value="Dicetak oleh Sistem MANPRO — Software Manajemen Proyek Terpadu")
    c.font = Font(name='Calibri', italic=True, size=9, color='999999')
    c.alignment = Alignment(horizontal='right', vertical='center')

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f'A1:F{row}'

    filename = f"Monthly_Report_{info.get('nama_proyek','')}_{bulan}.xlsx"
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ========================================================
# 5. TIME SCHEDULE & GANTT CHART (KURVA S)
# ========================================================
@app.route('/gantt/<int:proyek_id>')
@login_required
def gantt_chart(proyek_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # 1. Ambil detail proyek
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    proyek = cursor.fetchone()
    if not proyek:
        flash('Data proyek tidak ditemukan!', 'danger')
        return redirect(url_for('dashboard'))

    # 2. Ambil data rencana RAB / WBS
    cursor.execute("SELECT * FROM master_wbs WHERE proyek_id = %s ORDER BY id ASC", (proyek_id,))
    wbs_data = cursor.fetchall()

    # 3. Setup Variabel Kalkulasi Time Schedule
    minggu_total = 12 # Proyek dijadwalkan selesai dalam 12 minggu
    rencana_mingguan = [0] * minggu_total

    # Penjadwalan: Kapan mulai (minggu ke-) dan berapa lama durasinya (minggu)
    # Mapping Format: 'kode_wbs': (mulai_minggu, durasi_minggu)
    schedule_map = {
        '1.1': (1, 2),  # Persiapan: M1-M2
        '1.2': (2, 3),  # Pondasi: M2-M4
        '1.3': (3, 6),  # Dinding: M3-M8
        '1.4': (4, 5),  # Beton: M4-M8
        '1.5': (7, 3),  # Atap: M7-M9
        '1.6': (8, 3),  # Plafond: M8-M10
        '1.7': (9, 3),  # Lantai: M9-M11
        '1.8': (10, 3), # Landscape: M10-M12
        '1.9': (8, 4),  # Pintu & Jendela: M8-M11
        '1.10': (6, 6), # MEP: M6-M11
        '2.1': (1, 3), '2.2': (4, 4), '2.3': (8, 3), '2.4': (11, 2),
        '3.1': (1, 4), '3.2': (5, 3), '3.3': (8, 3), '3.4': (11, 2)
    }

    # Hitung bobot mingguan yang didistribusikan
    for item in wbs_data:
        kode = item.get('kode_wbs')
        bobot = float(item.get('bobot_persen', 0))

        # Simpan property durasi & mulai_minggu ke dict agar bisa dibaca di HTML
        item['start_week'] = 0
        item['duration'] = 0
        
        if kode in schedule_map:
            start_wk, duration = schedule_map[kode]
            item['start_week'] = start_wk
            item['duration'] = duration
            weekly_val = bobot / duration
            
            for w in range(start_wk, start_wk + duration):
                if 1 <= w <= minggu_total:
                    rencana_mingguan[w-1] += weekly_val

    # 4. Hitung Rencana Kumulatif (Kurva S Rencana)
    rencana_kumulatif = []
    cum_val = 0
    for val in rencana_mingguan:
        cum_val += val
        rencana_kumulatif.append(round(cum_val, 2))
        
    # Membulatkan nilai akhir proyek (koreksi presisi desimal agar 100%)
    if rencana_kumulatif and rencana_kumulatif[-1] > 99.0:
        rencana_kumulatif[-1] = 100.0

    # 5. Ambil Realisasi dari Laporan Harian (Kurva S Aktual)
    sql_kurva = """
        SELECT YEARWEEK(lh.tanggal_laporan, 1) as minggu,
               MAX(CAST(REPLACE(p.proses_kumulatif, '%', '') AS DECIMAL(5,2))) as progres
        FROM laporan_harian lh 
        JOIN pekerjaan p ON lh.id = p.laporan_id
        WHERE lh.proyek_id = %s 
        GROUP BY minggu 
        ORDER BY minggu ASC
    """
    cursor.execute(sql_kurva, (proyek_id,))
    data_harian = cursor.fetchall()
    
    labels = [f"Minggu {i+1}" for i in range(minggu_total)]
    
    # Proses realisasi disesuaikan dengan indeks minggu
    realisasi = []
    current_realisasi = 0
    for idx in range(len(data_harian)):
        try:
            current_realisasi = float(data_harian[idx]['progres'])
        except:
            pass
        realisasi.append(current_realisasi)

    conn.close()
    
    return render_template('gantt.html',
                            proyek=proyek,
                            labels=labels,
                            realisasi=realisasi,
                            rencana=rencana_kumulatif,
                            wbs_data=wbs_data,
                            rencana_mingguan=rencana_mingguan,
                            minggu_total=minggu_total)
# ==========================================
# 6. WBS MANAGEMENT
# ==========================================
@app.route('/wbs/<int:proyek_id>')
@login_required
def wbs_view(proyek_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    proyek = cursor.fetchone()
    cursor.execute("SELECT * FROM master_wbs WHERE proyek_id = %s ORDER BY kode_wbs ASC", (proyek_id,))
    wbs_list = cursor.fetchall()
    conn.close()
    return render_template('wbs.html', proyek=proyek, wbs_list=wbs_list)

@app.route('/add_wbs/<int:proyek_id>', methods=['POST'])
@login_required
def add_wbs(proyek_id):
    if not validate_csrf(): return redirect(url_for('wbs_view', proyek_id=proyek_id))
    # ── Permission Check ──
    if not can_user_edit(session['user_id']):
        flash('Anda tidak memiliki akses untuk menambah data!', 'danger')
        return redirect(url_for('wbs_view', proyek_id=proyek_id))
    conn = get_db_connection()
    cursor = conn.cursor()
    sql = "INSERT INTO master_wbs (proyek_id, kode_wbs, nama_pekerjaan, bobot_persen) VALUES (%s, %s, %s, %s)"
    cursor.execute(sql, (proyek_id, request.form['kode_wbs'], request.form['nama_pekerjaan'], request.form['bobot_persen']))
    conn.commit()
    cursor.close()
    conn.close()
    flash('Item WBS Berhasil Ditambahkan!', 'success')
    return redirect(url_for('wbs_view', proyek_id=proyek_id))

@app.route('/edit_wbs/<int:id>/<int:proyek_id>', methods=['POST'])
@role_required('admin', 'manager')
def edit_wbs(id, proyek_id):
    if not validate_csrf(): return redirect(url_for('wbs_view', proyek_id=proyek_id))
    # ── Permission Check ──
    if not can_user_edit(session['user_id']):
        flash('Anda tidak memiliki akses untuk mengedit data!', 'danger')
        return redirect(url_for('wbs_view', proyek_id=proyek_id))
    conn = get_db_connection()
    cursor = conn.cursor()
    sql = "UPDATE master_wbs SET kode_wbs = %s, nama_pekerjaan = %s, bobot_persen = %s WHERE id = %s"
    cursor.execute(sql, (request.form['kode_wbs'], request.form['nama_pekerjaan'], request.form['bobot_persen'], id))
    conn.commit()
    cursor.close()
    conn.close()
    flash('Item WBS Berhasil Diperbarui!', 'success')
    return redirect(url_for('wbs_view', proyek_id=proyek_id))

@app.route('/delete_wbs/<int:id>/<int:proyek_id>')
@role_required('admin', 'manager')
def delete_wbs(id, proyek_id):
    # ── Permission Check ──
    if not can_user_delete(session['user_id']):
        flash('Anda tidak memiliki akses untuk menghapus data!', 'danger')
        return redirect(url_for('wbs', proyek_id=proyek_id))
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM master_wbs WHERE id = %s", (id,))
    conn.commit()
    cursor.close()
    conn.close()
    flash('Item WBS Berhasil Dihapus!', 'danger')
    return redirect(url_for('wbs_view', proyek_id=proyek_id))

# ==========================================
# 7. COST & BUDGET MANAGEMENT & TRANSAKSI KAS
# ==========================================
@app.route('/budget/<int:proyek_id>')
@login_required
def budget_view(proyek_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    proyek = cursor.fetchone()
    cursor.execute("SELECT * FROM master_budget WHERE proyek_id = %s ORDER BY id DESC", (proyek_id,))
    budget_list = cursor.fetchall()
    
    cursor.execute("SELECT * FROM kategori_budget WHERE proyek_id = %s ORDER BY id ASC", (proyek_id,))
    kategori_list = cursor.fetchall()
    
    cursor.execute("SELECT COALESCE(SUM(anggaran_pagu), 0) AS total FROM kategori_budget WHERE proyek_id = %s", (proyek_id,))
    res_pagu = cursor.fetchone()
    total_pagu = res_pagu['total'] if res_pagu else 0
    
    total_realisasi = sum([b['realisasi_biaya'] for b in budget_list])
    sisa_budget = total_pagu - total_realisasi
    
    conn.close()
    return render_template('budget.html', proyek=proyek, budget_list=budget_list, total_pagu=total_pagu, total_realisasi=total_realisasi, sisa_budget=sisa_budget, kategori_list=kategori_list)

@app.route('/transaksi_kas/<int:proyek_id>')
@login_required
def transaksi_kas_view(proyek_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    proyek = cursor.fetchone()
    
    tgl_mulai = request.args.get('tgl_mulai', '')
    tgl_selesai = request.args.get('tgl_selesai', '')
    kategori_filter = request.args.get('kategori', '')
    
    query = "SELECT * FROM master_budget WHERE proyek_id = %s"
    params = [proyek_id]
    
    if tgl_mulai and tgl_selesai:
        query += " AND tanggal_transaksi BETWEEN %s AND %s"
        params.extend([tgl_mulai, tgl_selesai])
    elif tgl_mulai:
        query += " AND tanggal_transaksi >= %s"
        params.append(tgl_mulai)
    elif tgl_selesai:
        query += " AND tanggal_transaksi <= %s"
        params.append(tgl_selesai)
        
    if kategori_filter:
        query += " AND kategori = %s"
        params.append(kategori_filter)
        
    query += " ORDER BY id DESC"
    
    cursor.execute(query, tuple(params))
    transaksi_list = cursor.fetchall()
    
    cursor.execute("SELECT * FROM kategori_budget WHERE proyek_id = %s ORDER BY id ASC", (proyek_id,))
    kategori_list = cursor.fetchall()
    
    conn.close()
    return render_template('transaksi_kas.html', proyek=proyek, transaksi_list=transaksi_list, kategori_list=kategori_list, tgl_mulai=tgl_mulai, tgl_selesai=tgl_selesai, kategori_filter=kategori_filter)

@app.route('/add_budget/<int:proyek_id>', methods=['POST'])
@login_required
def add_budget(proyek_id):
    if not validate_csrf():
        return redirect(url_for('transaksi_kas_view', proyek_id=proyek_id))
    # ── Permission Check ──
    if not can_user_edit(session['user_id']):
        flash('Anda tidak memiliki akses untuk menambah data!', 'danger')
        return redirect(url_for('transaksi_kas_view', proyek_id=proyek_id))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    no_transaksi = request.form.get('no_transaksi', '-').strip()
    tanggal = request.form.get('tanggal_transaksi', '')
    pembuat = session.get('username', 'Admin')
    
    keterangan_list = request.form.getlist('keterangan[]')
    kategori_list = request.form.getlist('kategori[]')
    kuantitas_list = request.form.getlist('kuantitas[]')
    harga_satuan_list = request.form.getlist('harga_satuan[]')
    
    try:
        for i in range(len(keterangan_list)):
            nama_item = keterangan_list[i].strip()
            if not nama_item:
                continue
            
            kategori = kategori_list[i] if i < len(kategori_list) else ''
            qty = float(kuantitas_list[i]) if i < len(kuantitas_list) and kuantitas_list[i] else 1.0
            harga = float(harga_satuan_list[i]) if i < len(harga_satuan_list) and harga_satuan_list[i] else 0.0
            subtotal = qty * harga
            
            sql = """INSERT INTO master_budget 
                      (proyek_id, kategori, nama_item, anggaran_pagu, realisasi_biaya, no_transaksi, tanggal_transaksi, jenis_kas, pembuat, kuantitas, harga_satuan) 
                      VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            cursor.execute(sql, (proyek_id, kategori, nama_item, 0, subtotal, no_transaksi, tanggal, 'keluar', pembuat, qty, harga))
            
        conn.commit()
        flash('Transaksi Kas Berhasil Ditambahkan!', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Gagal menyimpan transaksi: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()
        
    return redirect(url_for('transaksi_kas_view', proyek_id=proyek_id))

@app.route('/edit_budget/<int:id>/<int:proyek_id>', methods=['POST'])
@role_required('admin', 'manager')
def edit_budget(id, proyek_id):
    if not validate_csrf():
        return redirect(url_for('transaksi_kas_view', proyek_id=proyek_id))
    # ── Permission Check ──
    if not can_user_edit(session['user_id']):
        flash('Anda tidak memiliki akses untuk mengedit data!', 'danger')
        return redirect(url_for('transaksi_kas_view', proyek_id=proyek_id))
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT no_transaksi FROM master_budget WHERE id = %s", (id,))
    old_trx = cursor.fetchone()
    old_no_transaksi = old_trx['no_transaksi'] if old_trx and old_trx.get('no_transaksi') else None
    
    no_transaksi = request.form.get('no_transaksi', '-').strip()
    tanggal = request.form.get('tanggal_transaksi', '')
    jenis = request.form.get('jenis', 'keluar')
    pembuat = session.get('username', 'Admin')
    
    keterangan_list = request.form.getlist('keterangan[]')
    kategori_list = request.form.getlist('kategori[]')
    kuantitas_list = request.form.getlist('kuantitas[]')
    harga_satuan_list = request.form.getlist('harga_satuan[]')
    
    try:
        if old_no_transaksi:
            cursor.execute("DELETE FROM master_budget WHERE proyek_id = %s AND (no_transaksi = %s OR no_transaksi = %s)", 
                           (proyek_id, old_no_transaksi, no_transaksi))
        else:
            cursor.execute("DELETE FROM master_budget WHERE id = %s", (id,))
        
        for i in range(len(keterangan_list)):
            nama_item = keterangan_list[i].strip()
            if not nama_item:
                continue
                
            kategori = kategori_list[i] if i < len(kategori_list) else ''
            qty = float(kuantitas_list[i]) if i < len(kuantitas_list) and kuantitas_list[i] else 1.0
            harga = float(harga_satuan_list[i]) if i < len(harga_satuan_list) and harga_satuan_list[i] else 0.0
            subtotal = qty * harga
            
            sql_insert = """INSERT INTO master_budget 
                            (proyek_id, kategori, nama_item, anggaran_pagu, realisasi_biaya, no_transaksi, tanggal_transaksi, jenis_kas, pembuat, kuantitas, harga_satuan) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            cursor.execute(sql_insert, (proyek_id, kategori, nama_item, 0, subtotal, no_transaksi, tanggal, jenis, pembuat, qty, harga))
            
        conn.commit()
        flash('Transaksi Kas Berhasil Diperbarui!', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Gagal memperbarui transaksi: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()
        
    return redirect(url_for('transaksi_kas_view', proyek_id=proyek_id))

@app.route('/delete_budget/<int:id>/<int:proyek_id>')
@role_required('admin', 'manager')
def delete_budget(id, proyek_id):
    # ── Permission Check ──
    if not can_user_delete(session['user_id']):
        flash('Anda tidak memiliki akses untuk menghapus data!', 'danger')
        return redirect(url_for('budget', proyek_id=proyek_id))
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT no_transaksi FROM master_budget WHERE id = %s", (id,))
    row = cursor.fetchone()
    
    if row and row.get('no_transaksi'):
        cursor.execute("DELETE FROM master_budget WHERE proyek_id = %s AND no_transaksi = %s", (proyek_id, row['no_transaksi']))
    else:
        cursor.execute("DELETE FROM master_budget WHERE id = %s", (id,))
        
    conn.commit()
    cursor.close()
    conn.close()
    
    flash('Transaksi Berhasil Dihapus!', 'danger')
    return redirect(url_for('transaksi_kas_view', proyek_id=proyek_id))

# ==========================================
# MANAJEMEN KATEGORI BIAYA PER PROYEK
# ==========================================
@app.route('/add_kategori_budget/<int:proyek_id>', methods=['POST'])
@role_required('admin', 'manager')
def add_kategori_budget(proyek_id):
    if not validate_csrf(): return redirect(url_for('pengaturan_proyek', proyek_id=proyek_id))
    nama_kategori = request.form.get('nama_kategori', '').strip().upper()
    if nama_kategori:
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("INSERT INTO kategori_budget (proyek_id, nama_kategori, anggaran_pagu) VALUES (%s, %s, 0)", (proyek_id, nama_kategori))
            conn.commit()
            flash('Kategori baru berhasil ditambahkan ke proyek ini!', 'success')
        except mysql.connector.IntegrityError:
            flash('Kategori tersebut sudah ada di proyek ini!', 'warning')
        finally:
            cursor.close()
            conn.close()
    return redirect(request.referrer or url_for('pengaturan_proyek', proyek_id=proyek_id))

@app.route('/edit_kategori_budget/<int:id>/<int:proyek_id>', methods=['POST'])
@role_required('admin', 'manager')
def edit_kategori_budget(id, proyek_id):
    if not validate_csrf(): return redirect(url_for('pengaturan_proyek', proyek_id=proyek_id))
    # ── Permission Check ──
    if not can_user_edit(session['user_id']):
        flash('Anda tidak memiliki akses untuk mengedit data!', 'danger')
        return redirect(url_for('pengaturan_proyek', proyek_id=proyek_id))
    nama_baru = request.form.get('nama_kategori', '').strip().upper()
    if nama_baru:
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("UPDATE kategori_budget SET nama_kategori = %s WHERE id = %s AND proyek_id = %s", (nama_baru, id, proyek_id))
            conn.commit()
            flash('Kategori biaya berhasil diperbarui!', 'success')
        except mysql.connector.IntegrityError:
            flash('Nama kategori tersebut sudah ada dalam proyek ini!', 'danger')
        finally:
            cursor.close()
            conn.close()
    return redirect(request.referrer or url_for('pengaturan_proyek', proyek_id=proyek_id))

@app.route('/delete_kategori_budget/<int:id>/<int:proyek_id>')
@role_required('admin', 'manager')
def delete_kategori_budget(id, proyek_id):
    # ── Permission Check ──
    if not can_user_delete(session['user_id']):
        flash('Anda tidak memiliki akses untuk menghapus data!', 'danger')
        return redirect(url_for('pengaturan_proyek', proyek_id=proyek_id))
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM kategori_budget WHERE id = %s AND proyek_id = %s", (id, proyek_id))
        conn.commit()
        flash('Kategori biaya berhasil dihapus dari proyek ini!', 'danger')
    except Exception as e:
        flash(f'Gagal menghapus kategori: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()
    return redirect(request.referrer or url_for('pengaturan_proyek', proyek_id=proyek_id))


# ==========================================
# REKAPITULASI BUDGET & CETAK REKAPITULASI
# ==========================================
@app.route('/laporan_proyek_budget/<int:proyek_id>')
@login_required
def laporan_proyek_budget(proyek_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    proyek = cursor.fetchone()
    
    cursor.execute("SELECT * FROM master_proyek ORDER BY id DESC")
    proyek_list = cursor.fetchall()
    
    cursor.execute("SELECT * FROM kategori_budget WHERE proyek_id = %s ORDER BY id ASC", (proyek_id,))
    kategori_list = cursor.fetchall()
    
    sql_rekap = """
        SELECT DATE_FORMAT(tanggal_transaksi, '%Y-%m') as periode,
               DATE_FORMAT(tanggal_transaksi, '%M') as bulan,
               DATE_FORMAT(tanggal_transaksi, '%Y') as tahun,
               kategori,
               SUM(realisasi_biaya) as subtotal
        FROM master_budget 
        WHERE proyek_id = %s AND tanggal_transaksi IS NOT NULL AND tanggal_transaksi != ''
        GROUP BY periode, kategori
        ORDER BY periode DESC
    """
    cursor.execute(sql_rekap, (proyek_id,))
    raw_rekap = cursor.fetchall()
    
    rekap_dict = {}
    for r in raw_rekap:
        periode = r['periode']
        if periode not in rekap_dict:
            rekap_dict[periode] = {
                'bulan': r['bulan'],
                'tahun': r['tahun'],
                'kategori': {k['nama_kategori']: 0 for k in kategori_list},
                'total': 0
            }
        if r['kategori'] in rekap_dict[periode]['kategori']:
            rekap_dict[periode]['kategori'][r['kategori']] = r['subtotal']
            
    for periode, data in rekap_dict.items():
        data['total'] = sum(data['kategori'].values())
        
    total_realisasi_per_kategori = {k['nama_kategori']: 0 for k in kategori_list}
    for data in rekap_dict.values():
        for kat_name, val in data['kategori'].items():
            total_realisasi_per_kategori[kat_name] += val

    total_anggaran_kategori = sum([k.get('anggaran_pagu', 0) or 0 for k in kategori_list])
    grand_total_realisasi = sum([data['total'] for data in rekap_dict.values()])

    conn.close()
    return render_template('laporan_proyek_budget.html', 
                           proyek=proyek, 
                           proyek_list=proyek_list, 
                           kategori_list=kategori_list, 
                           rekap_dict=rekap_dict,
                           total_realisasi_per_kategori=total_realisasi_per_kategori,
                           total_per_kategori=total_realisasi_per_kategori,
                           total_anggaran_kategori=total_anggaran_kategori,
                           grand_total_realisasi=grand_total_realisasi)

@app.route('/cetak_laporan_proyek_budget/<int:proyek_id>')
@login_required
def cetak_laporan_proyek_budget(proyek_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    proyek = cursor.fetchone()
    
    cursor.execute("SELECT * FROM kategori_budget WHERE proyek_id = %s ORDER BY id ASC", (proyek_id,))
    kategori_list = cursor.fetchall()
    
    sql_rekap = """
        SELECT DATE_FORMAT(tanggal_transaksi, '%Y-%m') as periode,
               DATE_FORMAT(tanggal_transaksi, '%M') as bulan,
               DATE_FORMAT(tanggal_transaksi, '%Y') as tahun,
               kategori,
               SUM(realisasi_biaya) as subtotal
        FROM master_budget 
        WHERE proyek_id = %s AND tanggal_transaksi IS NOT NULL AND tanggal_transaksi != ''
        GROUP BY periode, kategori
        ORDER BY periode DESC
    """
    cursor.execute(sql_rekap, (proyek_id,))
    raw_rekap = cursor.fetchall()
    
    rekap_dict = {}
    for r in raw_rekap:
        periode = r['periode']
        if periode not in rekap_dict:
            rekap_dict[periode] = {
                'bulan': r['bulan'], 
                'tahun': r['tahun'], 
                'kategori': {k['nama_kategori']: 0 for k in kategori_list}, 
                'total': 0
            }
        if r['kategori'] in rekap_dict[periode]['kategori']:
            rekap_dict[periode]['kategori'][r['kategori']] = r['subtotal']
            
    for periode, data in rekap_dict.items():
        data['total'] = sum(data['kategori'].values())
        
    total_realisasi_per_kategori = {k['nama_kategori']: 0 for k in kategori_list}
    for data in rekap_dict.values():
        for kat_name, val in data['kategori'].items():
            total_realisasi_per_kategori[kat_name] += val

    total_anggaran_kategori = sum([k.get('anggaran_pagu', 0) or 0 for k in kategori_list])
    grand_total_realisasi = sum([data['total'] for data in rekap_dict.values()])

    conn.close()
    
    return render_template('cetak_laporan_proyek_budget.html', 
                           proyek=proyek, 
                           kategori_list=kategori_list, 
                           rekap_dict=rekap_dict,
                           total_realisasi_per_kategori=total_realisasi_per_kategori,
                           total_per_kategori=total_realisasi_per_kategori,
                           total_anggaran_kategori=total_anggaran_kategori,
                           grand_total_realisasi=grand_total_realisasi)
# ========================================================
# KATEGORI BIAYA PROYEK
# ========================================================

def sync_kategori_biaya_proyek(proyek_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Daftar standar kategori biaya proyek sesuai permintaan
    standard_categories = [
        'STAFF',
        'UPAH MP',
        'MATERIAL',
        'MCU',
        'PERALATAN',
        'TRANSPORT & AKOMODASI',
        'JASA SUBCONT',
        'CONSUMABLE',
        'ENTERTAIN',
        'LAIN-LAIN'
    ]
    
    for kat in standard_categories:
        cursor.execute(
            "SELECT id FROM master_kategori_biaya WHERE proyek_id = %s AND nama_kategori = %s", 
            (proyek_id, kat)
        )
        exists = cursor.fetchone()
        
        if not exists:
            cursor.execute(
                "INSERT INTO master_kategori_biaya (proyek_id, nama_kategori) VALUES (%s, %s)", 
                (proyek_id, kat)
            )
            
    conn.commit()
    conn.close()

@app.route('/kategori-biaya/<int:proyek_id>')
@login_required
def daftar_kategori_biaya(proyek_id):
    # Sinkronkan data otomatis agar kategori lengkap
    sync_kategori_biaya_proyek(proyek_id)
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Ambil detail proyek
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    proyek = cursor.fetchone()
    if not proyek:
        flash('Data proyek tidak ditemukan!', 'danger')
        return redirect(url_for('dashboard'))
        
    # Ambil daftar kategori biaya untuk proyek ini
    cursor.execute("SELECT * FROM master_kategori_biaya WHERE proyek_id = %s ORDER BY id ASC", (proyek_id,))
    kategori_list = cursor.fetchall()
    
    conn.close()
    
    return render_template('kategori_biaya.html', proyek=proyek, kategori_list=kategori_list)

# ==========================================
# EXPORT TRANSAKSI KAS & LAINNYA
# ==========================================
@app.route('/cetak_transaksi_kas/<int:proyek_id>')
@login_required
def cetak_transaksi_kas(proyek_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    proyek = cursor.fetchone()
    
    tgl_mulai = request.args.get('tgl_mulai', '')
    tgl_selesai = request.args.get('tgl_selesai', '')
    kategori_filter = request.args.get('kategori', '')
    
    query = "SELECT * FROM master_budget WHERE proyek_id = %s"
    params = [proyek_id]
    
    if tgl_mulai and tgl_selesai:
        query += " AND tanggal_transaksi BETWEEN %s AND %s"
        params.extend([tgl_mulai, tgl_selesai])
    elif tgl_mulai:
        query += " AND tanggal_transaksi >= %s"
        params.append(tgl_mulai)
    elif tgl_selesai:
        query += " AND tanggal_transaksi <= %s"
        params.append(tgl_selesai)
        
    if kategori_filter:
        query += " AND kategori = %s"
        params.append(kategori_filter)
        
    query += " ORDER BY id DESC"
    
    cursor.execute(query, tuple(params))
    transaksi_list = cursor.fetchall()
    
    total_pemasukan = sum([t['realisasi_biaya'] for t in transaksi_list if t.get('jenis_kas') == 'masuk'])
    total_pengeluaran = sum([t['realisasi_biaya'] for t in transaksi_list if t.get('jenis_kas') != 'masuk'])
    saldo_akhir = total_pemasukan - total_pengeluaran
    
    conn.close()
    return render_template('cetak_transaksi_kas.html', 
                            proyek=proyek, 
                            transaksi_list=transaksi_list, 
                            tgl_mulai=tgl_mulai, 
                            tgl_selesai=tgl_selesai,
                            total_pemasukan=total_pemasukan,
                            total_pengeluaran=total_pengeluaran,
                            saldo_akhir=saldo_akhir)

@app.route('/export_transaksi_kas_excel/<int:proyek_id>')
@login_required
def export_transaksi_kas_excel(proyek_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    proyek = cursor.fetchone()
    tgl_mulai = request.args.get('tgl_mulai', '')
    tgl_selesai = request.args.get('tgl_selesai', '')
    kategori_filter = request.args.get('kategori', '')
    query = "SELECT * FROM master_budget WHERE proyek_id = %s"
    params = [proyek_id]
    if tgl_mulai and tgl_selesai:
        query += " AND tanggal_transaksi BETWEEN %s AND %s"
        params.extend([tgl_mulai, tgl_selesai])
    elif tgl_mulai:
        query += " AND tanggal_transaksi >= %s"
        params.append(tgl_mulai)
    elif tgl_selesai:
        query += " AND tanggal_transaksi <= %s"
        params.append(tgl_selesai)
    if kategori_filter:
        query += " AND kategori = %s"
        params.append(kategori_filter)
    query += " ORDER BY id DESC"
    cursor.execute(query, tuple(params))
    transaksi_list = cursor.fetchall()
    conn.close()
    total_pemasukan = sum(t['realisasi_biaya'] for t in transaksi_list if t.get('jenis_kas') == 'masuk')
    total_pengeluaran = sum(t['realisasi_biaya'] for t in transaksi_list if t.get('jenis_kas') != 'masuk')
    saldo_akhir = total_pemasukan - total_pengeluaran
    wb = Workbook()
    ws = wb.active
    ws.title = "Transaksi Kas"
    section_fill = PatternFill(start_color="DBE2EF", end_color="DBE2EF", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    money_fmt = '#,##0'
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value="LAPORAN TRANSAKSI KAS").font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value="Proyek:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=str(proyek.get('nama_proyek', '')))
    row += 1
    if tgl_mulai or tgl_selesai:
        ws.cell(row=row, column=1, value="Periode:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{tgl_mulai or '...'} s/d {tgl_selesai or '...'}")
        row += 1
    row += 1
    headers = ["No", "No. Transaksi", "Tanggal", "Jenis Kas", "Kategori", "Uraian / Item", "Jumlah (Rp)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = section_fill
        cell.border = thin_border
    row += 1
    for i, t in enumerate(transaksi_list, 1):
        jenis = "Kas Masuk" if t.get('jenis_kas') == 'masuk' else "Kas Keluar"
        vals = [
            i,
            str(t.get('no_transaksi', '-') or f"TRX-{t.get('id','')}"),
            str(t.get('tanggal_transaksi', '') or '-'),
            jenis,
            str(t.get('kategori', '') or '-'),
            str(t.get('nama_item', '') or '-'),
            float(t.get('realisasi_biaya', 0) or 0)
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = thin_border
            if c == 7:
                cell.number_format = money_fmt
        row += 1
    
    row += 1
    ws.cell(row=row, column=5, value="Total Pemasukan:").font = Font(bold=True)
    ws.cell(row=row, column=7, value=total_pemasukan).number_format = money_fmt
    ws.cell(row=row, column=7).font = Font(bold=True, color="008000")
    row += 1
    ws.cell(row=row, column=5, value="Total Pengeluaran:").font = Font(bold=True)
    ws.cell(row=row, column=7, value=total_pengeluaran).number_format = money_fmt
    ws.cell(row=row, column=7).font = Font(bold=True, color="FF0000")
    row += 1
    ws.cell(row=row, column=5, value="Saldo Akhir:").font = Font(bold=True, size=12)
    ws.cell(row=row, column=7, value=saldo_akhir).number_format = money_fmt
    ws.cell(row=row, column=7).font = Font(bold=True, size=12)
    for col in ws.columns:
        if isinstance(col[0], MergedCell):
            continue
        max_len = max(len(str(cell.value or '')) for cell in col if not isinstance(cell, MergedCell))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    filename = f"Transaksi_Kas_{proyek.get('nama_proyek','')}.xlsx"
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/anggaran_pagu/<int:proyek_id>')
@login_required
def anggaran_pagu_view(proyek_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM master_proyek WHERE id = %s", (proyek_id,))
    proyek = cursor.fetchone()
    
    sql = """
        SELECT kb.nama_kategori AS kategori,
               kb.anggaran_pagu AS anggaran_pagu
        FROM kategori_budget kb
        WHERE kb.proyek_id = %s
        ORDER BY kb.id ASC
    """
    cursor.execute(sql, (proyek_id,))
    pagu_list = cursor.fetchall()
    
    cursor.execute("SELECT * FROM kategori_budget WHERE proyek_id = %s ORDER BY id ASC", (proyek_id,))
    kategori_list = cursor.fetchall()
    conn.close()
    return render_template('anggaran_pagu.html', proyek=proyek, pagu_list=pagu_list, kategori_list=kategori_list)

@app.route('/edit_anggaran_pagu/<string:kategori>/<int:proyek_id>', methods=['POST'])
@role_required('admin')
def edit_anggaran_pagu(kategori, proyek_id):
    if not validate_csrf(): return redirect(url_for('anggaran_pagu_view', proyek_id=proyek_id))
    # ── Permission Check ──
    if not can_user_edit(session['user_id']):
        flash('Anda tidak memiliki akses untuk mengedit data!', 'danger')
        return redirect(url_for('anggaran_pagu_view', proyek_id=proyek_id))
    conn = get_db_connection()
    cursor = conn.cursor()
    anggaran_pagu = request.form['anggaran_pagu']
    
    cursor.execute("UPDATE kategori_budget SET anggaran_pagu = %s WHERE proyek_id = %s AND nama_kategori = %s", 
                   (anggaran_pagu, proyek_id, kategori))
        
    conn.commit()
    cursor.close()
    conn.close()
    flash('Anggaran Pagu Berhasil Diperbarui!', 'success')
    return redirect(url_for('anggaran_pagu_view', proyek_id=proyek_id))

@app.route('/edit_pagu/<int:proyek_id>', methods=['POST'])
@role_required('admin')
def edit_pagu(proyek_id):
    if not validate_csrf(): return redirect(url_for('budget_view', proyek_id=proyek_id))
    # ── Permission Check ──
    if not can_user_edit(session['user_id']):
        flash('Anda tidak memiliki akses untuk mengedit data!', 'danger')
        return redirect(url_for('budget_view', proyek_id=proyek_id))
    conn = get_db_connection()
    cursor = conn.cursor()
    pagu_baru = request.form['pagu_kontrak_total']
    cursor.execute("UPDATE master_proyek SET pagu_kontrak_total = %s WHERE id = %s", (pagu_baru, proyek_id))
    conn.commit()
    cursor.close()
    conn.close()
    flash('Total Pagu Kontrak Berhasil Diperbarui!', 'success')
    return redirect(url_for('budget_view', proyek_id=proyek_id))

@app.route('/delete_pagu/<int:proyek_id>')
@role_required('admin')
def delete_pagu(proyek_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE master_proyek SET pagu_kontrak_total = 0 WHERE id = %s", (proyek_id,))
    conn.commit()
    cursor.close()
    conn.close()
    flash('Total Pagu Kontrak Berhasil Dihapus (Di-reset ke 0)!', 'warning')
    return redirect(url_for('budget_view', proyek_id=proyek_id))


# ═══════════════════════════════════════════════════════════════════════════
# AUTO-SYNC: Background thread untuk sinkronisasi otomatis local ↔ TiDB Cloud
# ═══════════════════════════════════════════════════════════════════════════
import threading
import time

# Global sync state
_sync_state = {
    'last_sync': None,
    'status': 'idle',          # idle / syncing / error
    'last_direction': '',      # push / pull / merge
    'last_rows_synced': 0,
    'last_error': '',
    'sync_count': 0,
    'auto_sync_enabled': True,
    'interval_seconds': 300,   # 5 menit
}

def get_sync_state():
    """Return copy of sync state untuk ditampilkan di UI."""
    return dict(_sync_state)

app.jinja_env.globals['get_sync_state'] = get_sync_state


app.jinja_env.globals['wbs_list_global'] = []


def is_cloud_environment():
    """Deteksi apakah app berjalan di cloud (Railway/TiDB) atau lokal."""
    return bool(os.environ.get('DATABASE_URL') or os.environ.get('MYSQLURL'))

app.jinja_env.globals['is_cloud'] = is_cloud_environment


def _get_cloud_conn():
    """Koneksi ke TiDB Cloud (hardcoded untuk local sync)."""
    try:
        return mysql.connector.connect(
            host='gateway01.ap-southeast-1.prod.aws.tidbcloud.com',
            port=4000,
            user='vqpHmv4RwShLMQa.root',
            password='R2HE6uHS62HSOtMS',
            database='db_proyek',
            ssl_disabled=False,
            ssl_verify_cert=False,
            ssl_verify_identity=False,
            connect_timeout=10
        )
    except Exception as e:
        print(f'[AUTO-SYNC] Cloud connection error: {e}')
        return None


def _get_local_conn():
    """Koneksi ke MySQL lokal."""
    try:
        return mysql.connector.connect(
            host='127.0.0.1', port=3306, user='root',
            password='', database='db_proyek',
            connect_timeout=5
        )
    except Exception as e:
        print(f'[AUTO-SYNC] Local connection error: {e}')
        return None


def _get_table_columns(cursor, table_name):
    """Ambil nama kolom dari tabel."""
    cursor.execute(f'DESCRIBE {table_name}')
    return [row[0] for row in cursor.fetchall()]


def _get_natural_key_columns(table_name):
    """Ambil kolom natural key untuk tabel tertentu."""
    # Single natural key
    nk = NATURAL_KEYS.get(table_name)
    if nk:
        return [nk]
    # Composite key
    ck = COMPOSITE_KEYS.get(table_name)
    if ck:
        return ck
    return None


def _record_exists(cursor, table_name, key_cols, row):
    """Cek apakah record sudah ada berdasarkan natural key."""
    if not key_cols:
        return None
    
    conditions = ' AND '.join([f'`{k}` = %s' for k in key_cols])
    vals = [row.get(k) for k in key_cols]
    
    # Handle datetime
    for i, v in enumerate(vals):
        if hasattr(v, 'isoformat'):
            vals[i] = v.isoformat()
    
    try:
        cursor.execute(f'SELECT id FROM `{table_name}` WHERE {conditions} LIMIT 1', vals)
        result = cursor.fetchone()
        if result:
            return result[0] if isinstance(result, (tuple, list)) else result.get('id')
    except:
        pass
    return None


def _sync_table_push(local_conn, cloud_conn, table_name):
    """Push data lokal → cloud dengan natural key dedup."""
    local_cur = local_conn.cursor(dictionary=True)
    cloud_cur = cloud_conn.cursor(dictionary=True)
    
    # Ambil kolom dari cloud
    cloud_cur.execute(f'DESCRIBE {table_name}')
    cloud_cols = [row[0] for row in cloud_cur.fetchall()]
    if not cloud_cols:
        local_cur.close(); cloud_cur.close()
        return 0
    
    # Ambil semua data lokal
    local_cur.execute(f'SELECT * FROM {table_name}')
    rows = local_cur.fetchall()
    local_cur.close()
    if not rows:
        cloud_cur.close()
        return 0
    
    # Filter kolom yang ada di kedua sisi
    common_cols = [c for c in cloud_cols if c in rows[0]]
    if not common_cols:
        cloud_cur.close()
        return 0
    
    key_cols = _get_natural_key_columns(table_name)
    non_key_cols = [c for c in common_cols if c not in (key_cols or []) and c != 'id']
    
    count = 0
    for row in rows:
        try:
            # Cek apakah record sudah ada
            existing_id = _record_exists(cloud_cur, table_name, key_cols, row)
            
            if existing_id is not None and non_key_cols:
                # UPDATE: record sudah ada
                update_sql = ', '.join([f'`{c}` = %s' for c in non_key_cols])
                vals = [row.get(c) for c in non_key_cols]
                for i, v in enumerate(vals):
                    if hasattr(v, 'isoformat'):
                        vals[i] = v.isoformat()
                vals.append(existing_id)
                cloud_cur.execute(f'UPDATE `{table_name}` SET {update_sql} WHERE id = %s', vals)
            else:
                # INSERT: record belum ada
                cols_to_insert = [c for c in common_cols if c != 'id']
                cols_str = ', '.join([f'`{c}`' for c in cols_to_insert])
                placeholders = ', '.join(['%s'] * len(cols_to_insert))
                vals = [row.get(c) for c in cols_to_insert]
                for i, v in enumerate(vals):
                    if hasattr(v, 'isoformat'):
                        vals[i] = v.isoformat()
                cloud_cur.execute(f'INSERT INTO `{table_name}` ({cols_str}) VALUES ({placeholders})', vals)
            count += 1
        except Exception as e:
            pass
    
    cloud_conn.commit()
    cloud_cur.close()
    return count


def _sync_table_pull(local_conn, cloud_conn, table_name):
    """Pull data cloud → lokal dengan natural key dedup."""
    cloud_cur = cloud_conn.cursor(dictionary=True)
    local_cur = local_conn.cursor(dictionary=True)
    
    # Ambil kolom dari local
    local_cur.execute(f'DESCRIBE {table_name}')
    local_cols = [row[0] for row in local_cur.fetchall()]
    if not local_cols:
        cloud_cur.close(); local_cur.close()
        return 0
    
    # Ambil semua data cloud
    try:
        cloud_cur.execute(f'SELECT * FROM {table_name}')
        rows = cloud_cur.fetchall()
    except:
        cloud_cur.close(); local_cur.close()
        return 0
    cloud_cur.close()
    if not rows:
        local_cur.close()
        return 0
    
    # Filter kolom yang ada di kedua sisi
    common_cols = [c for c in local_cols if c in rows[0]]
    if not common_cols:
        local_cur.close()
        return 0
    
    key_cols = _get_natural_key_columns(table_name)
    non_key_cols = [c for c in common_cols if c not in (key_cols or []) and c != 'id']
    
    count = 0
    for row in rows:
        try:
            # Cek apakah record sudah ada
            existing_id = _record_exists(local_cur, table_name, key_cols, row)
            
            if existing_id is not None and non_key_cols:
                # UPDATE: record sudah ada
                update_sql = ', '.join([f'`{c}` = %s' for c in non_key_cols])
                vals = [row.get(c) for c in non_key_cols]
                for i, v in enumerate(vals):
                    if hasattr(v, 'isoformat'):
                        vals[i] = v.isoformat()
                vals.append(existing_id)
                local_cur.execute(f'UPDATE `{table_name}` SET {update_sql} WHERE id = %s', vals)
            else:
                # INSERT: record belum ada
                cols_to_insert = [c for c in common_cols if c != 'id']
                cols_str = ', '.join([f'`{c}`' for c in cols_to_insert])
                placeholders = ', '.join(['%s'] * len(cols_to_insert))
                vals = [row.get(c) for c in cols_to_insert]
                for i, v in enumerate(vals):
                    if hasattr(v, 'isoformat'):
                        vals[i] = v.isoformat()
                local_cur.execute(f'INSERT INTO `{table_name}` ({cols_str}) VALUES ({placeholders})', vals)
            count += 1
        except Exception as e:
            pass
    
    local_conn.commit()
    local_cur.close()
    return count


def _run_auto_sync(direction='merge'):
    """Jalankan sync otomatis di background thread."""
    global _sync_state
    
    if _sync_state['status'] == 'syncing':
        print('[AUTO-SYNC] Already syncing, skip.')
        return
    
    _sync_state['status'] = 'syncing'
    _sync_state['last_error'] = ''
    total_rows = 0
    
    try:
        local_conn = _get_local_conn()
        cloud_conn = _get_cloud_conn()
        
        if not local_conn:
            _sync_state['status'] = 'error'
            _sync_state['last_error'] = 'MySQL lokal tidak terhubung'
            return
        if not cloud_conn:
            _sync_state['status'] = 'error'
            _sync_state['last_error'] = 'TiDB Cloud tidak terhubung'
            return
        
        print(f'[AUTO-SYNC] Starting {direction}...')
        
        # Tabel yang di-sync (dalam urutan foreign key)
        tables = [
            'users', 'master_proyek', 'master_kategori_biaya',
            'laporan_harian', 'tenaga_kerja', 'peralatan', 'material',
            'pekerjaan', 'kondisi_lapangan', 'pengesahan',
            'master_wbs', 'kategori_budget', 'master_budget',
            'menu_permissions', 'settings', 'user_projects',
            'user_permissions', 'kategori_biaya'
        ]
        
        for table in tables:
            try:
                if direction in ('push', 'merge'):
                    n = _sync_table_push(local_conn, cloud_conn, table)
                    total_rows += n
                    if n > 0:
                        print(f'  [PUSH] {table}: {n} rows')
                if direction in ('pull', 'merge'):
                    n = _sync_table_pull(local_conn, cloud_conn, table)
                    total_rows += n
                    if n > 0:
                        print(f'  [PULL] {table}: {n} rows')
            except Exception as e:
                print(f'  [SKIP] {table}: {e}')
        
        local_conn.close()
        cloud_conn.close()
        
        _sync_state['status'] = 'idle'
        _sync_state['last_sync'] = time.strftime('%Y-%m-%d %H:%M:%S')
        _sync_state['last_direction'] = direction
        _sync_state['last_rows_synced'] = total_rows
        _sync_state['sync_count'] += 1
        
        print(f'[AUTO-SYNC] Done! {total_rows} rows synced ({direction})')
        
    except Exception as e:
        _sync_state['status'] = 'error'
        _sync_state['last_error'] = str(e)
        print(f'[AUTO-SYNC] ERROR: {e}')


def _auto_sync_thread():
    """Background thread yang jalan terus untuk auto-sync."""
    print('[AUTO-SYNC] Background thread started (interval: 5 min)')
    
    # Tunggu 30 detik setelah startup sebelum sync pertama
    time.sleep(30)
    
    while True:
        try:
            if _sync_state['auto_sync_enabled']:
                _run_auto_sync('merge')
            else:
                print('[AUTO-SYNC] Auto-sync disabled, sleeping...')
        except Exception as e:
            print(f'[AUTO-SYNC] Thread error: {e}')
        
        time.sleep(_sync_state['interval_seconds'])


# ═══════════════════════════════════════════════════════════════════════════
# Routes untuk Auto-Sync Management
# ═══════════════════════════════════════════════════════════════════════════

@app.route('/sync/auto/toggle', methods=['POST'])
@role_required('admin')
def toggle_auto_sync():
    """Toggle auto-sync on/off."""
    if not validate_csrf(): return redirect(url_for('sync_data'))
    _sync_state['auto_sync_enabled'] = not _sync_state['auto_sync_enabled']
    status = 'AKTIF ✅' if _sync_state['auto_sync_enabled'] else 'NONAKTIF ❌'
    flash(f'Auto-Sync sekarang: {status}', 'info')
    return redirect(url_for('sync_data'))


@app.route('/sync/auto/interval', methods=['POST'])
@role_required('admin')
def set_sync_interval():
    """Set interval auto-sync."""
    if not validate_csrf(): return redirect(url_for('sync_data'))
    try:
        minutes = int(request.form.get('interval', 5))
        minutes = max(1, min(60, minutes))  # 1-60 menit
        _sync_state['interval_seconds'] = minutes * 60
        flash(f'Interval auto-sync diatur ke {minutes} menit', 'success')
    except:
        flash('Interval tidak valid!', 'danger')
    return redirect(url_for('sync_data'))


@app.route('/sync/auto/now', methods=['POST'])
@role_required('admin')
def trigger_auto_sync():
    """Trigger manual sync via background thread."""
    if not validate_csrf(): return redirect(url_for('sync_data'))
    if is_cloud_environment():
        flash('Auto-sync hanya bisa dijalankan dari komputer lokal!', 'warning')
        return redirect(url_for('sync_data'))
    direction = request.form.get('direction', 'merge')
    t = threading.Thread(target=_run_auto_sync, args=(direction,), daemon=True)
    t.start()
    flash(f'Sync {direction} dimulai di background... Cek status dalam beberapa detik.', 'info')
    return redirect(url_for('sync_data'))


if __name__ == '__main__':
    # Start auto-sync background thread
    if not os.environ.get('DATABASE_URL') and not os.environ.get('MYSQLURL'):
        # Hanya jalan di lokal (bukan di Railway/TiDB Cloud)
        sync_thread = threading.Thread(target=_auto_sync_thread, daemon=True)
        sync_thread.start()
        print('[AUTO-SYNC] Background sync thread started!')
    
    port = int(os.environ.get('PORT', 5050))
    app.run(debug=False, host='0.0.0.0', port=port)